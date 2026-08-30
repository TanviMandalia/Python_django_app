import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


class ExportService:
    @staticmethod
    def create_excel_response(filename, sheet_title, headers, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="F5C518")
        alignment = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment

        for row in rows:
            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response

    @staticmethod
    def create_pdf_response(filename, title, table_data, col_widths=None, orientation='portrait'):
        buffer = io.BytesIO()
        pagesize = landscape(letter) if orientation == 'landscape' else letter
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            name='ClinicTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1A1A1A'),
            spaceAfter=12,
            alignment=1
        )

        elements = [
            Paragraph(f"<b>PhysioRehab Clinic — {title}</b>", title_style),
            Spacer(1, 10),
        ]

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A1A1A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#F5C518')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F9F9F9')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
        return response

