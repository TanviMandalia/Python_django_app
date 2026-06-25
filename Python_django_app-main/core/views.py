# ════════════════════════════════════════════════════════════
# core/views.py — TOP IMPORTS (replace your existing imports)
# ════════════════════════════════════════════════════════════

import json
import random
import logging
from datetime import datetime, time as dtime, timedelta  # ← time as dtime ADDED

from .models import Blog
from .forms import BlogForm
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.cache import cache  # ← ADDED
from django.db.models import Q, Sum
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.auth.hashers import check_password

from django.core.mail import send_mail
from django.conf import settings

from .models import (
    Appointment,
    Attendance,
    DailyTask,
    LeaveApplication,
    Message,
    Notification,
    PasswordResetOTP,
    Profile,
    SalaryRecord,
    SessionNote,
    StaffProfile,
    UserActivity,
    ClinicSettings,
    Blog,
    Review,
    Hospital,
    ClinicAdmin,
    SubscriptionPlan,
    HospitalSubscription,
    SupportTicket,
    SupportReply,
)
from .email_utils import (
    send_appointment_confirmation,
    send_appointment_status_update,
    send_admin_new_appointment_alert,
    send_otp_email,
    send_leave_status_email,
    send_salary_paid_email,
    send_task_assigned_email,
    send_staff_welcome_email,
)
from .notifications import (
    notify_appointment_booked,
    notify_appointment_status,
    notify_leave_decision,
    notify_task_assigned,
    notify_salary_paid,
)

logger = logging.getLogger(__name__)

# Shift definitions
MORNING_START = dtime(10, 0)
MORNING_END = dtime(13, 0)
EVENING_START = dtime(16, 0)
EVENING_END = dtime(20, 0)
LATE_GRACE = 15


# ─── HELPERS ─────────────────────────────────────────────────


def get_admin_user():
    return User.objects.filter(is_superuser=True).first()


def is_online(user):
    try:
        activity = UserActivity.objects.get(user=user)
        if not activity.last_seen:
            return False
        return activity.last_seen >= timezone.now() - timedelta(minutes=2)
    except UserActivity.DoesNotExist:
        return False


# ─── PUBLIC PAGES ────────────────────────────────────────────


def home(request):
    approved_reviews = Review.objects.filter(is_approved=True).order_by("-created_at")[
        :8
    ]
    return render(request, "home.html", {"approved_reviews": approved_reviews})


def about(request):
    return render(request, "about.html")


def services(request):
    return render(request, "services.html")


def contact(request):
    if request.method == "POST":
        messages.success(request, "✅ Message sent! We will contact you soon.")
        return redirect("/contact/")
    return render(request, "contact.html")


# =========================
# PUBLIC BLOG LIST
# =========================
def blog_list(request):
    category = request.GET.get("category", "")
    blogs = Blog.objects.order_by("-created_at")
    if category:
        blogs = blogs.filter(category=category)
    categories = Blog.CATEGORY_CHOICES
    return render(
        request,
        "blog.html",
        {"blogs": blogs, "categories": categories, "active_category": category},
    )


# =========================
# PUBLIC BLOG DETAIL
# =========================
def blog_detail(request, slug):
    blog = get_object_or_404(Blog, slug=slug)
    return render(request, "blog_detail.html", {"blog": blog})


# =========================
# ADMIN BLOG LIST
# =========================
def admin_blog_list(request):
    blogs = Blog.objects.order_by("-created_at")

    return render(request, "blog_list.html", {"blogs": blogs})


# =========================
# ADMIN ADD BLOG
# =========================
def admin_blog_add(request):
    if request.method == "POST":
        Blog.objects.create(
            title=request.POST.get("title"),
            excerpt=request.POST.get("excerpt", ""),
            category=request.POST.get("category", "General"),
            content=request.POST.get("content"),
            image=request.FILES.get("image") or None,
            before_image=request.FILES.get("before_image") or None,
            after_image=request.FILES.get("after_image") or None,
        )
        return redirect("admin_blog_list")
    return render(request, "blog_form.html", {"categories": Blog.CATEGORY_CHOICES})


# =========================
# ADMIN EDIT BLOG
# =========================
def admin_blog_edit(request, id):
    blog = get_object_or_404(Blog, id=id)
    if request.method == "POST":
        blog.title = request.POST.get("title")
        blog.excerpt = request.POST.get("excerpt", "")
        blog.category = request.POST.get("category", "General")
        blog.content = request.POST.get("content")
        if request.FILES.get("image"):
            blog.image = request.FILES.get("image")
        if request.FILES.get("before_image"):
            blog.before_image = request.FILES.get("before_image")
        if request.FILES.get("after_image"):
            blog.after_image = request.FILES.get("after_image")
        blog.save()
        return redirect("admin_blog_list")
    return render(
        request, "blog_form.html", {"blog": blog, "categories": Blog.CATEGORY_CHOICES}
    )


# =========================
# ADMIN DELETE BLOG
# =========================
def admin_blog_delete(request, id):
    blog = get_object_or_404(Blog, id=id)

    if request.method == "POST":
        blog.delete()
        return redirect("admin_blog_list")

    return render(request, "blog_delete.html", {"blog": blog})


# ──────────────────────────────────────────────────────────
# REVIEWS
# ──────────────────────────────────────────────────────────


@login_required
def submit_review(request):
    if request.method == "POST":
        name = (
            request.POST.get("reviewer_name", "").strip()
            or request.user.get_full_name()
            or request.user.username
        )
        title = request.POST.get("reviewer_title", "").strip()
        rating = int(request.POST.get("rating", 5))
        message = request.POST.get("message", "").strip()
        if message:
            Review.objects.create(
                patient=request.user,
                reviewer_name=name,
                reviewer_title=title,
                rating=max(1, min(5, rating)),
                message=message,
                is_approved=False,
                added_by_admin=False,
            )
            messages.success(
                request, "Thank you! Your review has been submitted for approval."
            )
        return redirect("client_dashboard")
    return redirect("client_dashboard")


@login_required
def admin_reviews(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    reviews = Review.objects.all()
    pending_count = reviews.filter(is_approved=False).count()
    return render(
        request,
        "admin_reviews.html",
        {"reviews": reviews, "pending_count": pending_count},
    )


@login_required
def admin_add_review(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    if request.method == "POST":
        Review.objects.create(
            reviewer_name=request.POST.get("reviewer_name", ""),
            reviewer_title=request.POST.get("reviewer_title", ""),
            rating=int(request.POST.get("rating", 5)),
            message=request.POST.get("message", ""),
            is_approved=True,
            added_by_admin=True,
        )
        messages.success(request, "Review added successfully.")
        return redirect("admin_reviews")
    return render(request, "admin_add_review.html")


@login_required
def toggle_review_approval(request, review_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    review = get_object_or_404(Review, id=review_id)
    review.is_approved = not review.is_approved
    review.save()
    return redirect("admin_reviews")


@login_required
def admin_delete_review(request, review_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    review = get_object_or_404(Review, id=review_id)
    review.delete()
    return redirect("admin_reviews")


# ─── AUTH ────────────────────────────────────────────────────


def login_view(request):
    if request.user.is_authenticated:
        return _redirect_by_role(request.user)

    if request.method == "POST":
        ip = _get_client_ip(request)
        cache_key = f"login_attempts_{ip}"
        attempts = cache.get(cache_key, 0)

        # Double-check lockout (middleware also checks)
        if attempts >= 5:
            remaining = cache.ttl(cache_key) // 60
            messages.error(
                request, f"🔒 Account locked. Try again in {remaining} minute(s)."
            )
            return render(request, "login.html")

        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)

        if user:
            # ✅ Success — clear failed attempts
            cache.delete(cache_key)
            login(request, user)
            return _redirect_by_role(user)
        else:
            # ❌ Failed — increment counter
            attempts += 1
            cache.set(cache_key, attempts, timeout=15 * 60)  # 15 min window
            remaining_attempts = 5 - attempts

            if remaining_attempts > 0:
                messages.error(
                    request,
                    f"❌ Invalid username or password. "
                    f"{remaining_attempts} attempt(s) remaining before lockout.",
                )
            else:
                messages.error(
                    request,
                    "🔒 Too many failed attempts. Account locked for 15 minutes.",
                )

    return render(request, "login.html")


def _get_client_ip(request):
    x_forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded:
        return x_forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "0.0.0.0")


def get_shift(t):
    """Return which shift a time belongs to."""
    if MORNING_START <= t <= MORNING_END:
        return "morning"
    if EVENING_START <= t <= EVENING_END:
        return "evening"
    return None


def is_late(t, shift):
    """Check if clock-in time is late beyond grace period."""
    if shift == "morning":
        grace = (
            datetime.combine(datetime.today(), MORNING_START)
            + timedelta(minutes=LATE_GRACE)
        ).time()
        return t > grace
    if shift == "evening":
        grace = (
            datetime.combine(datetime.today(), EVENING_START)
            + timedelta(minutes=LATE_GRACE)
        ).time()
        return t > grace
    return False


def _is_platform_superadmin(user):
    if not user.is_authenticated or not user.is_superuser:
        return False
    profile = getattr(user, "profile", None)
    return bool(profile and profile.is_platform_admin)


def _is_clinic_admin(user):
    if not user.is_authenticated or not user.is_superuser:
        return False
    return not _is_platform_superadmin(user)


def platform_superadmin_required(view_func):
    from functools import wraps

    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("login")
        if not _is_platform_superadmin(request.user):
            return redirect("home")
        return view_func(request, *args, **kwargs)

    return wrapper


def _redirect_by_role(user):
    if user.is_superuser:
        if _is_platform_superadmin(user):
            return redirect("super_admin_dashboard")
        return redirect("admin_dashboard")
    if user.is_staff or hasattr(user, "staff_profile"):
        return redirect("staff_dashboard")
    return redirect("client_dashboard")


def register_view(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "")
        confirm_password = request.POST.get("confirm_password", "")

        if password != confirm_password:
            messages.error(request, "❌ Passwords do not match!")
            return redirect("register")
        if User.objects.filter(username=username).exists():
            messages.error(request, "❌ Username already taken!")
            return redirect("register")
        if User.objects.filter(email=email).exists():
            messages.error(request, "❌ Email already registered!")
            return redirect("register")

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
        )
        messages.success(request, "✅ Account created! Please login.")
        return redirect("login")
    return render(request, "register.html")


def logout_view(request):
    logout(request)
    return redirect("home")


# ─── PROFILE ─────────────────────────────────────────────────


@login_required
def profile_view(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    return render(request, "profile.html", {"profile": profile})


@login_required
def edit_profile(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if request.method == "POST":
        request.user.first_name = request.POST.get("first_name", "").strip()
        request.user.last_name = request.POST.get("last_name", "").strip()
        request.user.email = request.POST.get("email", "").strip()
        request.user.save()

        profile.phone_number = request.POST.get("phone_number", "").strip()
        profile.gender = request.POST.get("gender", "")
        profile.date_of_birth = request.POST.get("date_of_birth") or None
        profile.blood_group = request.POST.get("blood_group", "")
        profile.emergency_contact = request.POST.get("emergency_contact", "").strip()
        profile.medical_notes = request.POST.get("medical_notes", "").strip()
        profile.address = request.POST.get("address", "").strip()
        if request.FILES.get("profile_photo"):
            profile.profile_photo = request.FILES["profile_photo"]
        profile.save()

        messages.success(request, "✅ Profile updated successfully.")
        return redirect("profile")
    return render(request, "edit_profile.html", {"profile": profile})


# ─── PASSWORD RESET (OTP FLOW) ───────────────────────────────


def request_otp(request):
    """Step 1 – user enters their email."""
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, "❌ No account found with that email.")
            return render(request, "request_otp.html")

        # Expire old OTPs
        PasswordResetOTP.objects.filter(user=user).delete()

        otp = str(random.randint(100000, 999999))
        PasswordResetOTP.objects.create(user=user, otp=otp)

        ok = send_otp_email(email, otp, purpose="Password Reset")
        if ok:
            messages.success(request, f"✅ OTP sent to {email}. Check your inbox.")
        else:
            messages.warning(
                request, "⚠️ Could not send email. Check server email config."
            )

        request.session["reset_user_id"] = user.id
        return redirect("verify_otp")

    return render(request, "request_otp.html")


def verify_otp(request):
    """Step 2 – user enters OTP."""
    user_id = request.session.get("reset_user_id")
    if not user_id:
        messages.error(request, "Session expired. Please start again.")
        return redirect("request_otp")

    if request.method == "POST":
        entered_otp = request.POST.get("otp", "").strip()
        try:
            user = User.objects.get(id=user_id)
            otp_obj = PasswordResetOTP.objects.filter(user=user).latest("created_at")
        except (User.DoesNotExist, PasswordResetOTP.DoesNotExist):
            messages.error(request, "❌ Invalid or expired OTP.")
            return redirect("request_otp")

        if otp_obj.is_expired():
            messages.error(request, "❌ OTP expired. Please request a new one.")
            return redirect("request_otp")

        if entered_otp == otp_obj.otp:
            request.session["otp_verified"] = True
            otp_obj.delete()
            return redirect("reset_password")

        messages.error(request, "❌ Incorrect OTP. Try again.")
    return render(request, "verify_otp.html")


def resend_otp(request):
    """Resend OTP to same email."""
    user_id = request.session.get("reset_user_id")
    if not user_id:
        return redirect("request_otp")
    try:
        user = User.objects.get(id=user_id)
        PasswordResetOTP.objects.filter(user=user).delete()
        otp = str(random.randint(100000, 999999))
        PasswordResetOTP.objects.create(user=user, otp=otp)
        ok = send_otp_email(user.email, otp, purpose="Password Reset")
        if ok:
            messages.success(request, "✅ New OTP sent!")
        else:
            messages.warning(request, "⚠️ Could not send email.")
    except User.DoesNotExist:
        messages.error(request, "User not found.")
    return redirect("verify_otp")


def reset_password(request):
    """
    Step 3 - Set new password after OTP verification.
    """

    if not request.session.get("otp_verified"):
        messages.error(request, "Please verify OTP first.")
        return redirect("request_otp")

    user_id = request.session.get("reset_user_id")

    if not user_id:
        messages.error(request, "Session expired.")
        return redirect("request_otp")

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, "User not found.")
        return redirect("request_otp")

    if request.method == "POST":
        password = request.POST.get("password", "").strip()
        confirm_password = request.POST.get("confirm_password", "").strip()

        # Empty fields
        if not password or not confirm_password:
            messages.error(request, "All fields are required.")
            return redirect("reset_password")

        # Password match check
        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect("reset_password")

        # Minimum length
        if len(password) < 8:
            messages.error(request, "Password must contain at least 8 characters.")
            return redirect("reset_password")

        # Same as old password check
        if user.check_password(password):
            messages.error(
                request, "New password cannot be the same as your current password."
            )
            return redirect("reset_password")

        # Save password
        user.set_password(password)
        user.save()

        # Clear reset session
        request.session.pop("otp_verified", None)
        request.session.pop("reset_user_id", None)

        messages.success(request, "Password reset successfully. Please login.")

        return redirect("login")

    return render(request, "reset_password.html")


@login_required
def change_password_request(request):
    """Logged-in user changes password via OTP sent to their email."""
    if request.method == "POST":
        otp = str(random.randint(100000, 999999))
        PasswordResetOTP.objects.filter(user=request.user).delete()
        PasswordResetOTP.objects.create(user=request.user, otp=otp)

        ok = send_otp_email(request.user.email, otp, purpose="Password Change")
        if ok:
            messages.success(request, f"✅ OTP sent to {request.user.email}")
        else:
            messages.warning(request, "⚠️ Email not sent – check server config.")

        request.session["reset_user_id"] = request.user.id
        request.session["otp_verified"] = False
        return redirect("verify_otp")
    return render(request, "change_password.html")


# ─── NOTIFICATIONS ───────────────────────────────────────────


@login_required
def notifications_view(request):
    notifs = Notification.objects.filter(recipient=request.user)
    return render(
        request,
        "notifications.html",
        {
            "notifications": notifs,
            "today": timezone.now().date(),
        },
    )


@login_required
def mark_notification_read(request, notif_id):
    Notification.objects.filter(id=notif_id, recipient=request.user).update(
        is_read=True
    )
    return JsonResponse({"status": "ok"})


@login_required
def mark_all_read(request):
    Notification.objects.filter(recipient=request.user, is_read=False).update(
        is_read=True
    )
    messages.success(request, "✅ All notifications marked as read.")
    return redirect("notifications")


# ─── PAYMENT ─────────────────────────────────────────────────

import razorpay
import hmac
import hashlib
import json as _json

from .models import PaymentRecord, ClinicPromo, ClinicSubscriptionPayment


def _get_patient_hospital(user):
    """Get the hospital linked to a patient via their appointments."""
    from .models import Hospital
    appt = Appointment.objects.filter(patient=user).select_related("hospital").order_by("-date").first()
    if appt and hasattr(appt, "hospital") and appt.hospital:
        return appt.hospital
    return None


def _get_patient_razorpay(user):
    """Return (client, key_id, key_secret) using the patient's clinic/doctor Razorpay account.
    Falls back to super-admin keys from settings if the clinic hasn't set theirs."""
    from .models import Hospital
    hospital = _get_patient_hospital(user)
    key_id = (hospital.razorpay_key_id if hospital and hospital.razorpay_key_id
              else getattr(settings, "RAZORPAY_KEY_ID", ""))
    key_secret = (hospital.razorpay_key_secret if hospital and hospital.razorpay_key_secret
                  else getattr(settings, "RAZORPAY_KEY_SECRET", ""))
    if not key_id or not key_secret:
        return None, key_id, key_secret
    return razorpay.Client(auth=(key_id, key_secret)), key_id, key_secret


@login_required
def payments(request):
    my_payments = PaymentRecord.objects.filter(patient=request.user).order_by("-created_at")
    upcoming_appointments = Appointment.objects.filter(
        patient=request.user, status__in=["pending", "confirmed"]
    ).order_by("date")
    clinic = ClinicSettings.objects.first()
    upi_id = getattr(settings, "CLINIC_UPI_ID", "dhvanipatalia@upi")
    _, razorpay_key_id, _ = _get_patient_razorpay(request.user)
    return render(
        request,
        "payments.html",
        {
            "my_payments": my_payments,
            "upcoming_appointments": upcoming_appointments,
            "clinic": clinic,
            "razorpay_key_id": razorpay_key_id,
            "upi_id": upi_id,
        },
    )


@login_required
@require_POST
def razorpay_create_order(request):
    """Create a Razorpay order — money goes to the doctor/clinic account."""
    amount_str = request.POST.get("amount", "500")
    appt_id = request.POST.get("appointment_id", "")
    try:
        amount_paise = int(float(amount_str) * 100)
    except (ValueError, TypeError):
        amount_paise = 50000

    client, key_id, _ = _get_patient_razorpay(request.user)
    if not client:
        return JsonResponse({"error": "Razorpay is not configured. Contact admin."}, status=400)

    try:
        order = client.order.create({
            "amount": amount_paise,
            "currency": "INR",
            "receipt": f"rcpt_{request.user.id}_{timezone.now().strftime('%Y%m%d%H%M%S')}",
            "notes": {
                "user_id": str(request.user.id),
                "appointment_id": appt_id,
                "patient_name": request.user.get_full_name() or request.user.username,
            },
        })
    except Exception as e:
        logger.error(f"Razorpay create order error: {e}")
        return JsonResponse({"error": str(e)}, status=400)

    record = PaymentRecord.objects.create(
        patient=request.user,
        appointment_id=appt_id if appt_id else None,
        amount=float(amount_str),
        method="razorpay",
        status="pending",
        razorpay_order_id=order["id"],
    )

    return JsonResponse({
        "order_id": order["id"],
        "amount": amount_paise,
        "currency": "INR",
        "key_id": key_id,
        "record_id": record.id,
        "patient_name": request.user.get_full_name() or request.user.username,
        "patient_email": request.user.email,
    })


@login_required
@require_POST
def razorpay_verify_payment(request):
    """Verify Razorpay payment signature and mark record as paid."""
    razorpay_order_id = request.POST.get("razorpay_order_id", "")
    razorpay_payment_id = request.POST.get("razorpay_payment_id", "")
    razorpay_signature = request.POST.get("razorpay_signature", "")
    record_id = request.POST.get("record_id", "")

    _, _, key_secret = _get_patient_razorpay(request.user)
    if not key_secret:
        messages.error(request, "⚠️ Payment verification failed — Razorpay not configured.")
        return redirect("payments")

    generated_sig = hmac.new(
        key_secret.encode("utf-8"),
        f"{razorpay_order_id}|{razorpay_payment_id}".encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    if generated_sig == razorpay_signature:
        PaymentRecord.objects.filter(id=record_id, patient=request.user).update(
            status="paid",
            razorpay_payment_id=razorpay_payment_id,
            razorpay_signature=razorpay_signature,
            transaction_id=razorpay_payment_id,
        )
        messages.success(request, "✅ Payment successful! Payment ID: " + razorpay_payment_id)
    else:
        PaymentRecord.objects.filter(id=record_id, patient=request.user).update(status="failed")
        messages.error(request, "❌ Payment verification failed. If money was deducted, contact the clinic.")

    return redirect("payments")


@csrf_exempt
def razorpay_webhook(request):
    """Razorpay webhook — verifies via HMAC and updates PaymentRecord status."""
    key_secret = getattr(settings, "RAZORPAY_KEY_SECRET", "")
    webhook_secret = getattr(settings, "RAZORPAY_WEBHOOK_SECRET", key_secret)
    payload = request.body
    sig_header = request.META.get("HTTP_X_RAZORPAY_SIGNATURE", "")
    if webhook_secret and sig_header:
        try:
            generated = hmac.new(
                webhook_secret.encode("utf-8"), payload, hashlib.sha256
            ).hexdigest()
            if generated != sig_header:
                return HttpResponse(status=400)
            data = _json.loads(payload)
            event = data.get("event", "")
            if event == "payment.captured":
                payment_id = data["payload"]["payment"]["entity"]["id"]
                order_id = data["payload"]["payment"]["entity"].get("order_id", "")
                if order_id:
                    PaymentRecord.objects.filter(razorpay_order_id=order_id).update(
                        status="paid",
                        razorpay_payment_id=payment_id,
                        transaction_id=payment_id,
                    )
        except Exception as e:
            logger.error(f"Razorpay webhook error: {e}")
            return HttpResponse(status=400)
    return HttpResponse(status=200)


@login_required
def record_cash_payment(request):
    if request.method != "POST":
        return redirect("payments")
    amount = request.POST.get("amount", "0")
    appt_id = request.POST.get("appointment_id", "")
    notes = request.POST.get("notes", "")
    PaymentRecord.objects.create(
        patient=request.user,
        appointment_id=appt_id if appt_id else None,
        amount=float(amount) if amount else 0,
        method="cash",
        status="paid",
        notes=notes,
        transaction_id=f"CASH-{timezone.now().strftime('%Y%m%d%H%M%S')}",
    )
    messages.success(request, "✅ Cash payment recorded.")
    return redirect("payments")


@login_required
def record_upi_payment(request):
    if request.method != "POST":
        return redirect("payments")
    amount = request.POST.get("amount", "0")
    appt_id = request.POST.get("appointment_id", "")
    txn_id = request.POST.get("transaction_id", "")
    PaymentRecord.objects.create(
        patient=request.user,
        appointment_id=appt_id if appt_id else None,
        amount=float(amount) if amount else 0,
        method="upi",
        status="pending",
        transaction_id=txn_id,
        notes=f"UPI Ref: {txn_id}",
    )
    messages.success(request, "✅ UPI payment submitted. The doctor will verify and confirm it shortly.")
    return redirect("payments")


@login_required
def admin_payments(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    all_payments = PaymentRecord.objects.select_related(
        "patient", "appointment"
    ).order_by("-created_at")
    total_paid = all_payments.filter(status="paid").aggregate(t=Sum("amount"))["t"] or 0
    pending_count = all_payments.filter(status="pending").count()
    pending_payments = all_payments.filter(status="pending")
    return render(
        request,
        "admin_payments.html",
        {
            "payments": all_payments,
            "total_paid": total_paid,
            "pending_count": pending_count,
            "pending_payments": pending_payments,
        },
    )


@login_required
@require_POST
def confirm_payment(request, payment_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    pay = get_object_or_404(PaymentRecord, id=payment_id)
    pay.status = "paid"
    if not pay.transaction_id:
        pay.transaction_id = f"CONF-{timezone.now().strftime('%Y%m%d%H%M%S')}"
    pay.save()
    Notification.objects.create(
        recipient=pay.patient,
        message=f"✅ Your payment of ₹{pay.amount} has been confirmed by the clinic. Thank you!",
        link="/payments/",
    )
    messages.success(request, f"✅ Payment of ₹{pay.amount} confirmed for {pay.patient.get_full_name() or pay.patient.username}.")
    return redirect("admin_payments")


@login_required
@require_POST
def reject_payment(request, payment_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    pay = get_object_or_404(PaymentRecord, id=payment_id)
    pay.status = "failed"
    pay.save()
    Notification.objects.create(
        recipient=pay.patient,
        message=f"❌ Your payment of ₹{pay.amount} could not be verified. Please contact the clinic or try again.",
        link="/payments/",
    )
    messages.warning(request, f"❌ Payment of ₹{pay.amount} rejected for {pay.patient.get_full_name() or pay.patient.username}.")
    return redirect("admin_payments")


# ─── CLIENT DASHBOARD ────────────────────────────────────────


@login_required
def client_dashboard(request):
    if request.user.is_superuser:
        return redirect("admin_dashboard")
    appointments = Appointment.objects.filter(patient=request.user)
    upcoming = appointments.filter(status__in=["pending", "confirmed"])
    completed = appointments.filter(status="completed")
    recent_notifs = Notification.objects.filter(recipient=request.user)[:5]
    has_reviewed = Review.objects.filter(patient=request.user).exists()
    approved_reviews = Review.objects.filter(is_approved=True).order_by("-created_at")[:6]
    my_payments = PaymentRecord.objects.filter(patient=request.user).order_by("-created_at")[:5]
    upcoming_appointments = upcoming.order_by("date")
    _, razorpay_key_id, _ = _get_patient_razorpay(request.user)
    return render(
        request,
        "client_dashboard.html",
        {
            "appointments": appointments,
            "upcoming": upcoming,
            "completed": completed,
            "total": appointments.count(),
            "upcoming_count": upcoming.count(),
            "completed_count": completed.count(),
            "recent_notifs": recent_notifs,
            "has_reviewed": has_reviewed,
            "approved_reviews": approved_reviews,
            "my_payments": my_payments,
            "upcoming_appointments": upcoming_appointments,
            "razorpay_key_id": razorpay_key_id,
        },
    )


# ─── ADMIN DASHBOARD ─────────────────────────────────────────


@login_required
def admin_dashboard(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    all_appointments = Appointment.objects.all()
    pending = all_appointments.filter(status="pending").count()
    confirmed = all_appointments.filter(status="confirmed").count()
    completed = all_appointments.filter(status="completed").count()
    total_users = User.objects.filter(is_superuser=False, is_staff=False).count()
    new_messages = Message.objects.filter(receiver=request.user, is_read=False).count()
    total_staff = StaffProfile.objects.count()
    blogs = Blog.objects.order_by("-created_at")[:6]
    pending_reviews = Review.objects.filter(is_approved=False).count()
    recent_reviews = Review.objects.filter(is_approved=True).order_by("-created_at")[:4]
    dismissed = request.session.get("dismissed_promos", [])
    active_promos = [
        p
        for p in ClinicPromo.objects.filter(is_active=True)
        if p.is_live and str(p.id) not in dismissed
    ]
    from django.db.models import Sum
    recent_payments = PaymentRecord.objects.order_by("-created_at")[:6]
    total_paid = PaymentRecord.objects.filter(status="paid").aggregate(t=Sum("amount"))["t"] or 0
    pending_payments_count = PaymentRecord.objects.filter(status="pending").count()
    return render(
        request,
        "admin_dashboard.html",
        {
            "total_users": total_users,
            "all_appointments": all_appointments,
            "total_appointments": all_appointments.count(),
            "pending": pending,
            "confirmed": confirmed,
            "completed": completed,
            "new_messages": new_messages,
            "total_staff": total_staff,
            "blogs": blogs,
            "pending_reviews": pending_reviews,
            "recent_reviews": recent_reviews,
            "active_promos": active_promos,
            "recent_payments": recent_payments,
            "total_paid": total_paid,
            "pending_payments_count": pending_payments_count,
        },
    )


# ─── APPOINTMENTS ────────────────────────────────────────────


@login_required
def book_appointment(request):
    if request.method == "POST":
        service = request.POST.get("service", "")
        date = request.POST.get("date", "")
        time = request.POST.get("time", "")
        notes = request.POST.get("notes", "")
        name = request.user.get_full_name() or request.user.username
        email = request.user.email
        phone = getattr(request.user, "profile", None)
        phone = phone.phone_number if phone else ""

        appt = Appointment.objects.create(
            patient=request.user,
            name=name,
            email=email,
            phone=phone,
            service=service,
            date=date,
            time=time,
            notes=notes,
            status="pending",
        )

        # Email patient + alert admin
        send_appointment_confirmation(appt)
        admin = get_admin_user()
        if admin:
            send_admin_new_appointment_alert(appt)
            notify_appointment_booked(appt, admin)

        messages.success(request, "✅ Appointment booked! Confirmation email sent.")
        return redirect("my_appointments")
    return render(request, "book_appointment.html")


@login_required
def my_appointments(request):
    appointments = Appointment.objects.filter(patient=request.user)
    return render(request, "my_appointments.html", {"appointments": appointments})


@login_required
def admin_appointments(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    all_appointments = Appointment.objects.all().order_by("-created_at")
    return render(
        request,
        "admin_appointments.html",
        {
            "all_appointments": all_appointments,
            "pending_count": all_appointments.filter(status="pending").count(),
            "confirmed_count": all_appointments.filter(status="confirmed").count(),
            "completed_count": all_appointments.filter(status="completed").count(),
            "cancelled_count": all_appointments.filter(status="cancelled").count(),
        },
    )


@login_required
def update_appointment(request, appt_id, status):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    appt = get_object_or_404(Appointment, id=appt_id)
    valid = ["pending", "confirmed", "completed", "cancelled"]
    if status not in valid:
        messages.error(request, "Invalid status.")
        return redirect("admin_appointments")

    appt.status = status
    appt.save()

    # Email + notification
    send_appointment_status_update(appt)
    notify_appointment_status(appt)

    messages.success(request, f"✅ Appointment {status}.")
    return redirect("admin_appointments")


# ─── ADMIN PATIENTS ──────────────────────────────────────────


@login_required
def admin_patients(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    patients = (
        User.objects.filter(is_superuser=False, is_staff=False)
        .order_by("-date_joined")
        .prefetch_related("profile", "appointments", "patient_notes")
    )
    patient_data = []
    for p in patients:
        try:
            profile = p.profile
        except Exception:
            profile = None
        appts = Appointment.objects.filter(patient=p).order_by("date")
        appt_count = appts.count()
        first_appt = appts.first()
        last_completed = appts.filter(status="completed").order_by("-date").first()
        latest_note = SessionNote.objects.filter(patient=p).order_by("-date").first()
        total_fees = appts.filter(consultation_fee__isnull=False).aggregate(
            total=Sum("consultation_fee")
        )["total"]
        phone = None
        if profile and profile.phone_number:
            phone = profile.phone_number
        elif first_appt and first_appt.phone:
            phone = first_appt.phone
        patient_data.append(
            {
                "user": p,
                "profile": profile,
                "appt_count": appt_count,
                "joining_date": first_appt.date if first_appt else p.date_joined.date(),
                "completion_date": last_completed.date if last_completed else None,
                "diagnosis": latest_note.diagnosis if latest_note else None,
                "treatment": latest_note.treatment if latest_note else None,
                "total_fees": total_fees,
                "phone": phone,
            }
        )
    return render(request, "admin_patients.html", {"patient_data": patient_data})


# ─── CHAT ────────────────────────────────────────────────────


@login_required
def client_chat(request):
    admin = get_admin_user()
    if not admin:
        return render(request, "client_chat.html", {"error": "No admin account found."})

    if request.method == "POST":
        content = request.POST.get("content", "").strip()
        if content:
            Message.objects.create(
                sender=request.user,
                receiver=admin,
                content=content,
                status=Message.STATUS_SENT,
            )
        return redirect("client_chat")

    msgs = Message.objects.filter(
        Q(sender=request.user, receiver=admin) | Q(sender=admin, receiver=request.user)
    ).order_by("created_at")

    Message.objects.filter(sender=admin, receiver=request.user, is_read=False).update(
        is_read=True, status=Message.STATUS_READ
    )

    admin_activity = UserActivity.objects.filter(user=admin).first()
    return render(
        request,
        "client_chat.html",
        {
            "messages_list": msgs,
            "admin": admin,
            "admin_online": is_online(admin),
            "admin_last_seen": admin_activity.last_seen if admin_activity else None,
        },
    )


@login_required
def admin_chat(request):
    if not request.user.is_superuser:
        return redirect("home")
    patients = User.objects.filter(is_superuser=False, is_staff=False)
    patient_chats = []
    for patient in patients:
        last_msg = (
            Message.objects.filter(
                Q(sender=patient, receiver=request.user)
                | Q(sender=request.user, receiver=patient)
            )
            .order_by("-created_at")
            .first()
        )
        unread = Message.objects.filter(
            sender=patient, receiver=request.user, is_read=False
        ).count()
        patient_chats.append(
            {"patient": patient, "last_msg": last_msg, "unread_count": unread}
        )
    patient_chats.sort(
        key=lambda x: x["last_msg"].created_at if x["last_msg"] else timezone.now(),
        reverse=True,
    )
    return render(request, "admin_chat.html", {"patient_chats": patient_chats})


@login_required
def admin_chat_detail(request, patient_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    patient = get_object_or_404(User, id=patient_id)

    if request.method == "POST":
        content = request.POST.get("content", "").strip()
        if content:
            Message.objects.create(
                sender=request.user,
                receiver=patient,
                content=content,
                status=Message.STATUS_SENT,
            )
        return redirect("admin_chat_detail", patient_id=patient_id)

    msgs = Message.objects.filter(
        Q(sender=patient, receiver=request.user)
        | Q(sender=request.user, receiver=patient)
    ).order_by("created_at")

    Message.objects.filter(sender=patient, receiver=request.user, is_read=False).update(
        is_read=True, status=Message.STATUS_READ
    )

    patient_activity = UserActivity.objects.filter(user=patient).first()
    return render(
        request,
        "admin_chat_detail.html",
        {
            "patient": patient,
            "messages_list": msgs,
            "patient_online": is_online(patient),
            "patient_last_seen": patient_activity.last_seen
            if patient_activity
            else None,
        },
    )


@login_required
def delete_message(request, message_id):
    message = get_object_or_404(Message, id=message_id, sender=request.user)
    redirect_url = (
        f"/admin-chat/{message.receiver.id}/" if request.user.is_superuser else "/chat/"
    )
    message.delete()
    return redirect(redirect_url)


# ─── TYPING INDICATORS ───────────────────────────────────────


@csrf_exempt
@login_required
def start_typing(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            receiver_id = data.get("receiver_id")
        except Exception:
            receiver_id = None
        obj, _ = UserActivity.objects.get_or_create(user=request.user)
        obj.is_typing = True
        obj.typing_to_id = receiver_id
        obj.last_seen = timezone.now()  # ← timestamp every keystroke
        obj.save(update_fields=["is_typing", "typing_to_id", "last_seen"])
        return JsonResponse({"status": "started"})
    return JsonResponse({"status": "invalid"}, status=400)


@csrf_exempt
@login_required
def stop_typing(request):
    if request.method == "POST":
        obj, _ = UserActivity.objects.get_or_create(user=request.user)
        obj.is_typing = False
        obj.typing_to_id = None
        obj.last_seen = timezone.now()
        obj.save(update_fields=["is_typing", "typing_to_id", "last_seen"])
        return JsonResponse({"status": "stopped"})
    return JsonResponse({"status": "invalid"}, status=400)


@login_required
def check_typing(request, user_id):
    obj = UserActivity.objects.filter(user_id=user_id).first()
    if not obj:
        return JsonResponse({"is_typing": False})

    # Auto-expire typing after 3 seconds of no update
    if obj.is_typing:
        from datetime import timedelta

        time_diff = timezone.now() - obj.last_seen
        if time_diff.seconds > 3:
            obj.is_typing = False
            obj.typing_to_id = None
            obj.save(update_fields=["is_typing", "typing_to_id"])
            return JsonResponse({"is_typing": False})

    is_typing = obj.is_typing and obj.typing_to_id == request.user.id
    return JsonResponse({"is_typing": is_typing})


# ─── ADMIN – STAFF ───────────────────────────────────────────


@login_required
def admin_staff(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    staff_list = StaffProfile.objects.all().select_related("user")
    return render(request, "admin_staff.html", {"staff_list": staff_list})


@login_required
def add_staff(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "")
        role = request.POST.get("role", "")
        phone = request.POST.get("phone", "")
        salary = request.POST.get("salary", 0)

        # ── Validation checks ──
        if not username:
            messages.error(request, "❌ Username is required.")
            return redirect("add_staff")

        if User.objects.filter(username=username).exists():
            messages.error(
                request,
                f'❌ Username "{username}" is already taken. Please choose a different one.',
            )
            return redirect("add_staff")

        if User.objects.filter(email=email).exists():
            messages.error(request, f'❌ Email "{email}" is already registered.')
            return redirect("add_staff")

        if len(password) < 6:
            messages.error(request, "❌ Password must be at least 6 characters.")
            return redirect("add_staff")

        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_staff=True,
            )
            StaffProfile.objects.create(
                user=user, role=role, phone=phone, salary=salary
            )
            send_staff_welcome_email(user, password)
            messages.success(
                request,
                f"✅ Staff member {first_name} {last_name} added successfully! Welcome email sent.",
            )
            return redirect("admin_staff")

        except Exception as e:
            messages.error(request, f"❌ Error creating staff: {str(e)}")
            return redirect("add_staff")

    return render(request, "add_staff.html")


# ─── ADMIN – LEAVES ──────────────────────────────────────────


@login_required
def admin_leaves(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")

    # ── Auto-approve any expired pending leaves ──
    from django.utils.timezone import localdate

    today = localdate()
    expired = LeaveApplication.objects.filter(status="pending", to_date__lt=today)
    for leave in expired:
        leave.status = "approved"
        leave.admin_note = "Auto-approved: Admin did not respond before leave date."
        leave.save()
        # Email staff
        try:
            from django.core.mail import send_mail

            send_mail(
                subject="✅ Leave Auto-Approved — No Admin Response",
                message=f"""
Dear {leave.staff.get_full_name() or leave.staff.username},

Your leave has been AUTO-APPROVED as the admin did not respond before the leave date.

  Leave Type : {leave.get_leave_type_display()}
  From       : {leave.from_date.strftime("%d %B %Y")}
  To         : {leave.to_date.strftime("%d %B %Y")}

Regards,
Dr. Dhvani Patalia Physio Rehab
                """.strip(),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[leave.staff.email],
                fail_silently=True,
            )
        except Exception:
            pass

    # existing code continues below ──
    leaves = LeaveApplication.objects.all().select_related("staff")
    pending = leaves.filter(status="pending").count()
    return render(request, "admin_leaves.html", {"leaves": leaves, "pending": pending})


@login_required
def update_leave(request, leave_id, status):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    leave = get_object_or_404(LeaveApplication, id=leave_id)
    if status in ["approved", "rejected"]:
        leave.status = status
        leave.admin_note = request.POST.get("admin_note", "")
        leave.save()

        # Email + notification
        send_leave_status_email(leave)
        notify_leave_decision(leave)

        messages.success(request, f"✅ Leave {status}.")
    return redirect("admin_leaves")


# ─── ADMIN – ATTENDANCE ──────────────────────────────────────


@login_required
def admin_attendance(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    today = timezone.now().date()
    today_att = Attendance.objects.filter(date=today).select_related("staff")
    all_attendance = Attendance.objects.all().select_related("staff")[:50]
    return render(
        request,
        "admin_attendance.html",
        {
            "today_attendance": today_att,
            "all_attendance": all_attendance,
            "today": today,
        },
    )


# ─── ADMIN – SALARY ──────────────────────────────────────────


@login_required
def admin_salary(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    if request.method == "POST":
        staff_id = request.POST.get("staff_id")
        month = request.POST.get("month")
        year = request.POST.get("year")
        basic = request.POST.get("basic_salary", 0)
        bonus = request.POST.get("bonus", 0)
        deduction = request.POST.get("deduction", 0)
        net = float(basic) + float(bonus) - float(deduction)

        staff_user = get_object_or_404(User, id=staff_id)
        record = SalaryRecord.objects.create(
            staff=staff_user,
            month=month,
            year=year,
            basic_salary=basic,
            bonus=bonus,
            deduction=deduction,
            net_salary=net,
            is_paid=True,
            paid_on=timezone.now().date(),
        )

        # Email + notification
        send_salary_paid_email(record)
        notify_salary_paid(record)

        messages.success(request, "✅ Salary record saved and email sent!")
        return redirect("admin_salary")

    salary_records = SalaryRecord.objects.all().select_related("staff")
    staff_list = StaffProfile.objects.all().select_related("user")
    return render(
        request,
        "admin_salary.html",
        {
            "salary_records": salary_records,
            "staff_list": staff_list,
        },
    )


# ─── ADMIN – TASKS ───────────────────────────────────────────


@login_required
def admin_tasks(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    tasks = DailyTask.objects.all().select_related("assigned_to", "assigned_by")
    return render(request, "admin_tasks.html", {"tasks": tasks})


@login_required
def add_task(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    if request.method == "POST":
        staff_id = request.POST.get("staff_id")
        title = request.POST.get("title", "")
        description = request.POST.get("description", "")
        priority = request.POST.get("priority", "medium")
        due_date = request.POST.get("due_date") or None

        staff_user = get_object_or_404(User, id=staff_id)
        task = DailyTask.objects.create(
            assigned_to=staff_user,
            assigned_by=request.user,
            title=title,
            description=description,
            priority=priority,
            due_date=due_date,
        )

        # Email + notification
        send_task_assigned_email(task)
        notify_task_assigned(task)

        messages.success(
            request, f"✅ Task assigned to {staff_user.get_full_name()}! Email sent."
        )
        return redirect("admin_tasks")

    staff_list = StaffProfile.objects.all().select_related("user")
    return render(request, "add_task.html", {"staff_list": staff_list})


# ─── ADMIN – SESSION NOTES ───────────────────────────────────


@login_required
def admin_session_notes(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    notes = SessionNote.objects.all().select_related("staff", "patient")
    return render(request, "admin_session_notes.html", {"notes": notes})


# ─── STAFF DASHBOARD ─────────────────────────────────────────


def _is_staff(user):
    return user.is_staff or hasattr(user, "staff_profile")


@login_required
def staff_dashboard(request):
    if not _is_staff(request.user):
        return redirect("client_dashboard")
    today = timezone.now().date()
    today_att = Attendance.objects.filter(staff=request.user, date=today).first()
    pending_tasks = DailyTask.objects.filter(
        assigned_to=request.user, status="pending"
    ).count()
    pending_leaves = LeaveApplication.objects.filter(
        staff=request.user, status="pending"
    ).count()
    my_tasks = DailyTask.objects.filter(assigned_to=request.user).order_by(
        "-created_at"
    )[:5]
    return render(
        request,
        "staff_dashboard.html",
        {
            "today_attendance": today_att,
            "pending_tasks": pending_tasks,
            "pending_leaves": pending_leaves,
            "my_tasks": my_tasks,
            "today": today,
        },
    )


# ════════════════════════════════════════════════════════════
# FILE: core/views.py
# FIND your existing staff_attendance function and
# REPLACE the ENTIRE function with this complete version
# ════════════════════════════════════════════════════════════


@login_required
def staff_attendance(request):
    if not _is_staff(request.user):
        return redirect("client_dashboard")

    today = timezone.now().date()
    now_dt = timezone.localtime(timezone.now())
    now_time = now_dt.time()

    today_record = Attendance.objects.filter(staff=request.user, date=today).first()

    # ── Shift window flags ──
    in_morning = MORNING_START <= now_time <= MORNING_END
    in_evening = EVENING_START <= now_time <= EVENING_END
    before_morning = now_time < MORNING_START
    between_shifts = MORNING_END < now_time < EVENING_START
    after_evening = now_time > EVENING_END

    if request.method == "POST":
        action = request.POST.get("action")

        # ── MORNING CLOCK IN ────────────────────────────────
        if action == "morning_clock_in":
            if not in_morning:
                messages.error(
                    request,
                    f"❌ Morning shift is 10:00 AM – 1:00 PM only. "
                    f"Current time: {now_time.strftime('%I:%M %p')}",
                )
                return redirect("staff_attendance")

            if today_record and today_record.clock_in:
                messages.warning(request, "⚠️ Already clocked in for morning shift.")
                return redirect("staff_attendance")

            record, _ = Attendance.objects.get_or_create(staff=request.user, date=today)
            record.clock_in = now_time
            record.notes = ""
            record.save()

            if is_late(now_time, "morning"):
                late_mins = int(
                    (
                        datetime.combine(today, now_time)
                        - datetime.combine(today, MORNING_START)
                    ).seconds
                    / 60
                )
                record.notes = f"Late morning clock-in by {late_mins} min."
                record.save()
                _send_late_email(request.user, "Morning", now_time, late_mins)
                messages.warning(
                    request,
                    f"⚠️ Clocked in late for morning shift by {late_mins} minutes. "
                    f"Email notification sent.",
                )
            else:
                messages.success(
                    request, f"✅ Morning clock-in at {now_time.strftime('%I:%M %p')}"
                )

        # ── MORNING CLOCK OUT ───────────────────────────────
        elif action == "morning_clock_out":
            if not today_record or not today_record.clock_in:
                messages.error(
                    request, "❌ You have not clocked in for morning shift yet."
                )
                return redirect("staff_attendance")

            if today_record.morning_clock_out:
                messages.warning(request, "⚠️ Already clocked out from morning shift.")
                return redirect("staff_attendance")

            today_record.morning_clock_out = now_time

            # Calculate morning hours
            ci = datetime.combine(today, today_record.clock_in)
            co = datetime.combine(today, now_time)
            morning_hours = max(round((co - ci).seconds / 3600, 2), 0)
            today_record.morning_hours = morning_hours

            # Check early leave
            if now_time < MORNING_END:
                short_mins = int(
                    (
                        datetime.combine(today, MORNING_END)
                        - datetime.combine(today, now_time)
                    ).seconds
                    / 60
                )
                note = f"Left morning shift {short_mins} min early."
                today_record.notes = (today_record.notes + " " + note).strip()
                _send_early_leave_email(request.user, "Morning", now_time, short_mins)
                messages.warning(
                    request,
                    f"⚠️ Clocked out {short_mins} min early from morning shift. "
                    f"Email notification sent.",
                )
            else:
                messages.success(
                    request,
                    f"✅ Morning clock-out at {now_time.strftime('%I:%M %p')} "
                    f"— {morning_hours} hrs",
                )
            today_record.save()

        # ── EVENING CLOCK IN ────────────────────────────────
        elif action == "evening_clock_in":
            if not in_evening:
                messages.error(
                    request,
                    f"❌ Evening shift is 4:00 PM – 8:00 PM only. "
                    f"Current time: {now_time.strftime('%I:%M %p')}",
                )
                return redirect("staff_attendance")

            if today_record and today_record.evening_clock_in:
                messages.warning(request, "⚠️ Already clocked in for evening shift.")
                return redirect("staff_attendance")

            record, _ = Attendance.objects.get_or_create(staff=request.user, date=today)

            # ── If morning was fully missed → First Half Leave ──
            if not record.clock_in and not record.morning_clock_out:
                if "First Half Leave" not in record.notes:
                    record.notes = (
                        record.notes + " First Half Leave (Morning shift missed)."
                    ).strip()
                    LeaveApplication.objects.get_or_create(
                        staff=request.user,
                        from_date=today,
                        to_date=today,
                        defaults={
                            "leave_type": "casual",
                            "reason": "Auto-marked: First Half Leave (morning shift missed)",
                            "status": "pending",
                        },
                    )
                    _send_half_day_email(request.user, "Morning")
                    messages.warning(
                        request,
                        "⚠️ Morning shift was missed. Marked as First Half Leave. "
                        "Leave application auto-submitted for admin approval.",
                    )

            record.evening_clock_in = now_time
            record.save()
            today_record = record

            if is_late(now_time, "evening"):
                late_mins = int(
                    (
                        datetime.combine(today, now_time)
                        - datetime.combine(today, EVENING_START)
                    ).seconds
                    / 60
                )
                today_record.notes = (
                    today_record.notes + f" Late evening clock-in by {late_mins} min."
                ).strip()
                today_record.save()
                _send_late_email(request.user, "Evening", now_time, late_mins)
                messages.warning(
                    request,
                    f"⚠️ Clocked in late for evening shift by {late_mins} minutes. "
                    f"Email notification sent.",
                )
            else:
                if "First Half Leave" not in (today_record.notes or ""):
                    messages.success(
                        request,
                        f"✅ Evening clock-in at {now_time.strftime('%I:%M %p')}",
                    )

        # ── EVENING CLOCK OUT ───────────────────────────────
        elif action == "evening_clock_out":
            if not today_record or not today_record.evening_clock_in:
                messages.error(
                    request, "❌ You have not clocked in for evening shift yet."
                )
                return redirect("staff_attendance")

            if today_record.clock_out:
                messages.warning(request, "⚠️ Already clocked out from evening shift.")
                return redirect("staff_attendance")

            today_record.clock_out = now_time

            # Calculate evening hours
            ci = datetime.combine(today, today_record.evening_clock_in)
            co = datetime.combine(today, now_time)
            evening_hours = max(round((co - ci).seconds / 3600, 2), 0)
            today_record.evening_hours = evening_hours

            # Total = morning + evening
            morning_h = float(today_record.morning_hours or 0)
            today_record.total_hours = round(morning_h + evening_hours, 2)

            # Check early leave
            if now_time < EVENING_END:
                short_mins = int(
                    (
                        datetime.combine(today, EVENING_END)
                        - datetime.combine(today, now_time)
                    ).seconds
                    / 60
                )
                note = f"Left evening shift {short_mins} min early."
                today_record.notes = (today_record.notes + " " + note).strip()
                _send_early_leave_email(request.user, "Evening", now_time, short_mins)
                messages.warning(
                    request,
                    f"⚠️ Clocked out {short_mins} min early from evening shift. "
                    f"Email notification sent.",
                )
            else:
                messages.success(
                    request,
                    f"✅ Day complete! Total: {today_record.total_hours} hrs 🎉",
                )

            # ── If evening missed before this point was caught by cron ──
            # Mark Second Half Leave if morning done but evening was missed
            # (This handles edge case where staff clocks out very late)
            today_record.save()

        else:
            messages.error(request, "❌ Invalid action.")

        return redirect("staff_attendance")

    # ── GET request ─────────────────────────────────────────
    all_att = Attendance.objects.filter(staff=request.user).order_by("-date")[:30]

    return render(
        request,
        "staff_attendance.html",
        {
            "today_record": today_record,
            "all_attendance": all_att,
            "today": today,
            "now_time": now_time,
            "in_morning": in_morning,
            "in_evening": in_evening,
            "before_morning": before_morning,
            "between_shifts": between_shifts,
            "after_evening": after_evening,
            "MORNING_START": MORNING_START.strftime("%I:%M %p"),
            "MORNING_END": MORNING_END.strftime("%I:%M %p"),
            "EVENING_START": EVENING_START.strftime("%I:%M %p"),
            "EVENING_END": EVENING_END.strftime("%I:%M %p"),
        },
    )


# ════════════════════════════════════════════════════════════
# ADD these 3 helper functions anywhere in core/views.py
# AFTER the staff_attendance function
# ════════════════════════════════════════════════════════════


def _send_late_email(user, shift_name, clock_time, late_mins):
    """Send late arrival email to staff."""
    # from django.core.mail import send_mail
    # from django.conf import settings
    try:
        send_mail(
            subject=f"⚠️ Late Attendance — {shift_name} Shift",
            message=f"""
Dear {user.get_full_name() or user.username},

You clocked in LATE for the {shift_name} Shift today.

  Clock-in Time : {clock_time.strftime("%I:%M %p")}
  Late By       : {late_mins} minutes
  Date          : {timezone.now().date().strftime("%d %B %Y")}

Shift Timings:
  Morning  — 10:00 AM to 1:00 PM
  Evening  — 4:00 PM to 8:00 PM

Regards,
Dr. Dhvani Patalia Physio Rehab
            """.strip(),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=True,
        )
    except Exception:
        pass


def _send_early_leave_email(user, shift_name, clock_time, short_mins):
    """Send early leave email to staff."""
    from django.core.mail import send_mail
    from django.conf import settings

    try:
        send_mail(
            subject=f"⚠️ Early Leave Detected — {shift_name} Shift",
            message=f"""
Dear {user.get_full_name() or user.username},

You clocked out EARLY from the {shift_name} Shift today.

  Clock-out Time : {clock_time.strftime("%I:%M %p")}
  Left Early By  : {short_mins} minutes
  Date           : {timezone.now().date().strftime("%d %B %Y")}

If you had a valid reason, please inform the admin.

Regards,
Dr. Dhvani Patalia Physio Rehab
            """.strip(),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=True,
        )
    except Exception:
        pass


def _send_half_day_email(user, missed_shift):
    """Send half day leave email to staff."""
    from django.core.mail import send_mail
    from django.conf import settings

    half = "First" if missed_shift == "Morning" else "Second"
    timing = "10:00 AM – 1:00 PM" if missed_shift == "Morning" else "4:00 PM – 8:00 PM"
    try:
        send_mail(
            subject=f"📋 {half} Half Leave Auto-Marked — {missed_shift} Shift Missed",
            message=f"""
Dear {user.get_full_name() or user.username},

Since you missed the {missed_shift} shift today, your attendance
has been automatically marked as {half.upper()} HALF LEAVE.

  Date          : {timezone.now().date().strftime("%d %B %Y")}
  Missed Shift  : {missed_shift} ({timing})
  Status        : {half} Half Leave (Pending Admin Approval)

A leave application has been auto-submitted on your behalf.
If this was an error, please contact admin to correct it.

Regards,
Dr. Dhvani Patalia Physio Rehab
            """.strip(),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=True,
        )
    except Exception:
        pass


@login_required
def staff_leave(request):
    if not _is_staff(request.user):
        return redirect("client_dashboard")
    if request.method == "POST":
        LeaveApplication.objects.create(
            staff=request.user,
            leave_type=request.POST.get("leave_type"),
            from_date=request.POST.get("from_date"),
            to_date=request.POST.get("to_date"),
            reason=request.POST.get("reason", ""),
        )
        messages.success(request, "✅ Leave application submitted!")
        return redirect("staff_leave")
    my_leaves = LeaveApplication.objects.filter(staff=request.user)
    return render(request, "staff_leave.html", {"my_leaves": my_leaves})


@login_required
def staff_salary(request):
    if not _is_staff(request.user):
        return redirect("client_dashboard")
    salary_records = SalaryRecord.objects.filter(staff=request.user)
    total_earned = sum(s.net_salary for s in salary_records if s.is_paid)
    return render(
        request,
        "staff_salary.html",
        {
            "salary_records": salary_records,
            "total_earned": total_earned,
        },
    )


@login_required
def staff_tasks(request):
    if not _is_staff(request.user):
        return redirect("client_dashboard")
    tasks = DailyTask.objects.filter(assigned_to=request.user)
    return render(
        request,
        "staff_tasks.html",
        {
            "tasks": tasks,
            "pending": tasks.filter(status="pending").count(),
            "in_progress": tasks.filter(status="in_progress").count(),
            "completed": tasks.filter(status="completed").count(),
        },
    )


@login_required
def update_task(request, task_id, status):
    task = get_object_or_404(DailyTask, id=task_id, assigned_to=request.user)
    if status in ["pending", "in_progress", "completed"]:
        task.status = status
        if status == "completed":
            task.completed_at = timezone.now()
        task.save()
        messages.success(request, f"✅ Task marked as {status}.")
    return redirect("staff_tasks")


@login_required
def staff_session_notes(request):
    if not _is_staff(request.user):
        return redirect("client_dashboard")
    notes = SessionNote.objects.filter(staff=request.user)
    return render(request, "staff_session_notes.html", {"notes": notes})


@login_required
def add_session_note(request):
    if not _is_staff(request.user):
        return redirect("client_dashboard")
    if request.method == "POST":
        patient = get_object_or_404(User, id=request.POST.get("patient_id"))
        SessionNote.objects.create(
            staff=request.user,
            patient=patient,
            diagnosis=request.POST.get("diagnosis", ""),
            treatment=request.POST.get("treatment", ""),
            next_session=request.POST.get("next_session", ""),
        )
        messages.success(request, "✅ Session note added!")
        return redirect("staff_session_notes")
    patients = User.objects.filter(is_superuser=False, is_staff=False)
    return render(request, "add_session_note.html", {"patients": patients})


# ─── ADMIN – APPOINTMENTS CRUD ───────────────────────────────


@login_required
def add_appointment(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    patients = User.objects.filter(is_superuser=False, is_staff=False)
    if request.method == "POST":
        patient_id = request.POST.get("patient_id")
        service = request.POST.get("service", "")
        date = request.POST.get("date", "")
        time = request.POST.get("time", "")
        notes = request.POST.get("notes", "")
        status = request.POST.get("status", "pending")
        patient = get_object_or_404(User, id=patient_id)
        Appointment.objects.create(
            patient=patient,
            name=patient.get_full_name() or patient.username,
            email=patient.email,
            phone=getattr(getattr(patient, "profile", None), "phone_number", ""),
            service=service,
            date=date,
            time=time,
            notes=notes,
            status=status,
        )
        messages.success(request, "✅ Appointment added successfully!")
        return redirect("admin_appointments")
    return render(request, "add_appointment.html", {"patients": patients})


@login_required
def edit_appointment(request, appt_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    appt = get_object_or_404(Appointment, id=appt_id)
    if request.method == "POST":
        appt.service = request.POST.get("service", appt.service)
        appt.date = request.POST.get("date", str(appt.date))
        appt.time = request.POST.get("time", appt.time)
        appt.notes = request.POST.get("notes", appt.notes)
        appt.status = request.POST.get("status", appt.status)
        appt.save()
        messages.success(request, "✅ Appointment updated successfully!")
        return redirect("admin_appointments")
    return render(request, "edit_appointment.html", {"appt": appt})


@login_required
def delete_appointment(request, appt_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    appt = get_object_or_404(Appointment, id=appt_id)
    if request.method == "POST":
        appt.delete()
        messages.success(request, "✅ Appointment deleted.")
        return redirect("admin_appointments")
    return render(
        request,
        "confirm_delete.html",
        {
            "title": "Delete Appointment",
            "item": f"{appt.name} — {appt.get_service_display()} on {appt.date}",
            "cancel_url": "/admin-appointments/",
        },
    )


# ─── ADMIN – PATIENTS CRUD ───────────────────────────────────


@login_required
def edit_patient(request, user_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    patient = get_object_or_404(User, id=user_id)
    profile, _ = Profile.objects.get_or_create(user=patient)
    if request.method == "POST":
        patient.first_name = request.POST.get("first_name", "").strip()
        patient.last_name = request.POST.get("last_name", "").strip()
        patient.email = request.POST.get("email", "").strip()
        patient.save()
        profile.phone_number = request.POST.get("phone_number", "").strip()
        profile.address = request.POST.get("address", "").strip()
        profile.save()
        messages.success(request, "✅ Patient updated successfully!")
        return redirect("admin_patients")
    return render(
        request, "edit_patient.html", {"patient": patient, "profile": profile}
    )


@login_required
def delete_patient(request, user_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    patient = get_object_or_404(User, id=user_id, is_superuser=False, is_staff=False)
    if request.method == "POST":
        patient.is_active = False
        patient.save()
        messages.success(
            request,
            f'✅ Patient "{patient.get_full_name() or patient.username}" has been deactivated.',
        )
        return redirect("admin_patients")
    return render(
        request,
        "confirm_delete.html",
        {
            "title": "Deactivate Patient",
            "item": patient.get_full_name() or patient.username,
            "message": "This will deactivate the patient account. They will not be able to log in. You can reactivate them anytime.",
            "cancel_url": "/admin-patients/",
        },
    )


@login_required
def reactivate_patient(request, user_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    patient = get_object_or_404(User, id=user_id, is_superuser=False, is_staff=False)
    patient.is_active = True
    patient.save()
    messages.success(
        request,
        f'✅ Patient "{patient.get_full_name() or patient.username}" has been reactivated.',
    )
    return redirect("admin_patients")


# ─── ADMIN – STAFF CRUD ──────────────────────────────────────


@login_required
def edit_staff(request, staff_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    staff = get_object_or_404(StaffProfile, id=staff_id)
    if request.method == "POST":
        staff.user.first_name = request.POST.get("first_name", "").strip()
        staff.user.last_name = request.POST.get("last_name", "").strip()
        staff.user.email = request.POST.get("email", "").strip()
        staff.user.save()
        staff.role = request.POST.get("role", staff.role)
        staff.phone = request.POST.get("phone", "").strip()
        staff.salary = request.POST.get("salary", staff.salary)
        staff.save()
        messages.success(request, "✅ Staff member updated successfully!")
        return redirect("admin_staff")
    return render(request, "edit_staff.html", {"staff": staff})


@login_required
def delete_staff(request, staff_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    staff = get_object_or_404(StaffProfile, id=staff_id)
    if request.method == "POST":
        staff.user.delete()
        messages.success(request, "✅ Staff member deleted.")
        return redirect("admin_staff")
    return render(
        request,
        "confirm_delete.html",
        {
            "title": "Delete Staff Member",
            "item": staff.user.get_full_name() or staff.user.username,
            "cancel_url": "/admin-staff/",
        },
    )


# ─── ADMIN – ATTENDANCE CRUD ─────────────────────────────────


@login_required
def add_attendance(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    staff_list = StaffProfile.objects.all().select_related("user")
    if request.method == "POST":
        staff_user = get_object_or_404(User, id=request.POST.get("staff_id"))
        date = request.POST.get("date")
        clock_in = request.POST.get("clock_in") or None
        clock_out = request.POST.get("clock_out") or None
        total_hours = request.POST.get("total_hours", 0)
        notes = request.POST.get("notes", "")
        Attendance.objects.update_or_create(
            staff=staff_user,
            date=date,
            defaults={
                "clock_in": clock_in,
                "clock_out": clock_out,
                "total_hours": total_hours,
                "notes": notes,
            },
        )
        messages.success(request, "✅ Attendance record saved!")
        return redirect("admin_attendance")
    return render(request, "add_attendance.html", {"staff_list": staff_list})


@login_required
def edit_attendance(request, att_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    att = get_object_or_404(Attendance, id=att_id)
    if request.method == "POST":
        att.clock_in = request.POST.get("clock_in") or None
        att.clock_out = request.POST.get("clock_out") or None
        att.total_hours = request.POST.get("total_hours", att.total_hours)
        att.notes = request.POST.get("notes", "")
        att.save()
        messages.success(request, "✅ Attendance updated!")
        return redirect("admin_attendance")
    return render(request, "edit_attendance.html", {"att": att})


@login_required
def delete_attendance(request, att_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    att = get_object_or_404(Attendance, id=att_id)
    if request.method == "POST":
        att.delete()
        messages.success(request, "✅ Attendance record deleted.")
        return redirect("admin_attendance")
    return render(
        request,
        "confirm_delete.html",
        {
            "title": "Delete Attendance Record",
            "item": f"{att.staff.get_full_name() or att.staff.username} — {att.date}",
            "cancel_url": "/admin-attendance/",
        },
    )


# ─── ADMIN – LEAVES CRUD ─────────────────────────────────────


@login_required
def delete_leave(request, leave_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    leave = get_object_or_404(LeaveApplication, id=leave_id)
    if request.method == "POST":
        leave.delete()
        messages.success(request, "✅ Leave application deleted.")
        return redirect("admin_leaves")
    return render(
        request,
        "confirm_delete.html",
        {
            "title": "Delete Leave Application",
            "item": f"{leave.staff.get_full_name() or leave.staff.username} — {leave.from_date} to {leave.to_date}",
            "cancel_url": "/admin-leaves/",
        },
    )


# ─── ADMIN – SALARY CRUD ─────────────────────────────────────


@login_required
def edit_salary(request, record_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    record = get_object_or_404(SalaryRecord, id=record_id)
    if request.method == "POST":
        record.month = request.POST.get("month", record.month)
        record.year = request.POST.get("year", record.year)
        record.basic_salary = request.POST.get("basic_salary", record.basic_salary)
        record.bonus = request.POST.get("bonus", record.bonus)
        record.deduction = request.POST.get("deduction", record.deduction)
        record.net_salary = (
            float(record.basic_salary) + float(record.bonus) - float(record.deduction)
        )
        record.save()
        messages.success(request, "✅ Salary record updated!")
        return redirect("admin_salary")
    return render(request, "edit_salary.html", {"record": record})


@login_required
def delete_salary(request, record_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    record = get_object_or_404(SalaryRecord, id=record_id)
    if request.method == "POST":
        record.delete()
        messages.success(request, "✅ Salary record deleted.")
        return redirect("admin_salary")
    return render(
        request,
        "confirm_delete.html",
        {
            "title": "Delete Salary Record",
            "item": f"{record.staff.get_full_name() or record.staff.username} — {record.month} {record.year}",
            "cancel_url": "/admin-salary/",
        },
    )


# ─── ADMIN – TASKS CRUD ──────────────────────────────────────


@login_required
def edit_task_admin(request, task_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    task = get_object_or_404(DailyTask, id=task_id)
    staff_list = StaffProfile.objects.all().select_related("user")
    if request.method == "POST":
        task.title = request.POST.get("title", task.title)
        task.description = request.POST.get("description", "")
        task.priority = request.POST.get("priority", task.priority)
        task.status = request.POST.get("status", task.status)
        task.due_date = request.POST.get("due_date") or None
        new_staff_id = request.POST.get("staff_id")
        if new_staff_id:
            task.assigned_to = get_object_or_404(User, id=new_staff_id)
        task.save()
        messages.success(request, "✅ Task updated!")
        return redirect("admin_tasks")
    return render(request, "edit_task.html", {"task": task, "staff_list": staff_list})


@login_required
def delete_task_admin(request, task_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    task = get_object_or_404(DailyTask, id=task_id)
    if request.method == "POST":
        task.delete()
        messages.success(request, "✅ Task deleted.")
        return redirect("admin_tasks")
    return render(
        request,
        "confirm_delete.html",
        {
            "title": "Delete Task",
            "item": task.title,
            "cancel_url": "/admin-tasks/",
        },
    )


# ─── ADMIN – SESSION NOTES CRUD ──────────────────────────────


@login_required
def admin_add_session_note(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    patients = User.objects.filter(is_superuser=False, is_staff=False)
    staff_list = StaffProfile.objects.all().select_related("user")
    if request.method == "POST":
        staff = get_object_or_404(User, id=request.POST.get("staff_id"))
        patient = get_object_or_404(User, id=request.POST.get("patient_id"))
        SessionNote.objects.create(
            staff=staff,
            patient=patient,
            diagnosis=request.POST.get("diagnosis", ""),
            treatment=request.POST.get("treatment", ""),
            next_session=request.POST.get("next_session", ""),
        )
        messages.success(request, "✅ Session note added!")
        return redirect("admin_session_notes")
    return render(
        request,
        "admin_add_session_note.html",
        {"patients": patients, "staff_list": staff_list},
    )


@login_required
def edit_session_note(request, note_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    note = get_object_or_404(SessionNote, id=note_id)
    patients = User.objects.filter(is_superuser=False, is_staff=False)
    staff_list = StaffProfile.objects.all().select_related("user")
    if request.method == "POST":
        note.staff = get_object_or_404(User, id=request.POST.get("staff_id"))
        note.patient = get_object_or_404(User, id=request.POST.get("patient_id"))
        note.diagnosis = request.POST.get("diagnosis", "")
        note.treatment = request.POST.get("treatment", "")
        note.next_session = request.POST.get("next_session", "")
        note.save()
        messages.success(request, "✅ Session note updated!")
        return redirect("admin_session_notes")
    return render(
        request,
        "edit_session_note.html",
        {"note": note, "patients": patients, "staff_list": staff_list},
    )


@login_required
def delete_session_note(request, note_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    note = get_object_or_404(SessionNote, id=note_id)
    if request.method == "POST":
        note.delete()
        messages.success(request, "✅ Session note deleted.")
        return redirect("admin_session_notes")
    return render(
        request,
        "confirm_delete.html",
        {
            "title": "Delete Session Note",
            "item": f"{note.patient.get_full_name() or note.patient.username} — {note.date}",
            "cancel_url": "/admin-session-notes/",
        },
    )


@login_required
def admin_settings(request):
    settings_obj, created = ClinicSettings.objects.get_or_create(id=1)

    if request.method == "POST":
        settings_obj.clinic_name = request.POST.get("clinic_name")

        settings_obj.phone = request.POST.get("phone")

        settings_obj.email = request.POST.get("email")

        settings_obj.address = request.POST.get("address")

        if request.FILES.get("logo"):
            settings_obj.logo = request.FILES["logo"]

        settings_obj.save()

        messages.success(request, "Settings updated successfully.")

        return redirect("admin_settings")

    return render(request, "admin_settings.html", {"settings": settings_obj})


# ════════════════════════════════════════════════════════════
# Add these two views to your core/views.py
# ════════════════════════════════════════════════════════════


@login_required
def progress_tracking(request):
    if request.user.is_superuser:
        # Admin: can view any patient's progress
        all_patients = User.objects.filter(is_superuser=False, is_staff=False)
        patient_id = request.GET.get("patient_id")

        if patient_id:
            selected_patient = get_object_or_404(User, id=patient_id)
            appointments = Appointment.objects.filter(patient=selected_patient)
            session_notes = SessionNote.objects.filter(
                patient=selected_patient
            ).select_related("staff")
        else:
            selected_patient = None
            appointments = Appointment.objects.all()
            session_notes = SessionNote.objects.all().select_related("staff", "patient")
    else:
        # Client: sees only their own data
        all_patients = None
        selected_patient = None
        appointments = Appointment.objects.filter(patient=request.user)
        session_notes = SessionNote.objects.filter(patient=request.user).select_related(
            "staff"
        )

    # ── Stats ──
    total_appointments = appointments.count()
    completed_count = appointments.filter(status="completed").count()
    upcoming_count = appointments.filter(status__in=["pending", "confirmed"]).count()
    session_notes_count = session_notes.count()

    # ── Recovery percentage ──
    recovery_pct = round(
        (completed_count / total_appointments * 100) if total_appointments else 0
    )
    attendance_pct = round(
        (
            (total_appointments - appointments.filter(status="cancelled").count())
            / total_appointments
            * 100
        )
        if total_appointments
        else 0
    )
    consistency_pct = min(recovery_pct + 10, 100)
    notes_pct = round(
        (session_notes_count / completed_count * 100) if completed_count else 0
    )

    # ── Service breakdown ──
    service_map = {
        "orthopedic": "Orthopedic",
        "neurological": "Neurological",
        "sports": "Sports",
        "pediatric": "Pediatric",
        "womens": "Women's Health",
        "home_visit": "Home Visit",
    }
    service_breakdown = []
    for key, label in service_map.items():
        count = appointments.filter(service=key).count()
        if count > 0:
            service_breakdown.append(
                {
                    "label": label,
                    "count": count,
                    "pct": round(count / total_appointments * 100)
                    if total_appointments
                    else 0,
                }
            )
    service_breakdown.sort(key=lambda x: x["count"], reverse=True)

    return render(
        request,
        "progress_tracking.html",
        {
            "appointments": appointments,
            "session_notes": session_notes,
            "all_patients": all_patients,
            "selected_patient": selected_patient,
            "total_appointments": total_appointments,
            "completed_count": completed_count,
            "upcoming_count": upcoming_count,
            "session_notes_count": session_notes_count,
            "recovery_pct": recovery_pct,
            "attendance_pct": attendance_pct,
            "consistency_pct": consistency_pct,
            "notes_pct": notes_pct,
            "service_breakdown": service_breakdown,
        },
    )


@login_required
def reports_analytics(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")

    all_appointments = Appointment.objects.all()
    total_appointments = all_appointments.count()
    completed_appointments = all_appointments.filter(status="completed").count()
    confirmed_appointments = all_appointments.filter(status="confirmed").count()
    pending_appointments = all_appointments.filter(status="pending").count()
    cancelled_appointments = all_appointments.filter(status="cancelled").count()

    # ── Donut chart percentages ──
    def pct(n):
        return round(n / total_appointments * 100) if total_appointments else 0

    confirmed_pct = pct(confirmed_appointments)
    completed_pct = pct(completed_appointments)
    pending_pct = pct(pending_appointments)

    # ── Service bar chart ──
    service_map = [
        ("orthopedic", "Ortho"),
        ("neurological", "Neuro"),
        ("sports", "Sports"),
        ("pediatric", "Pedia"),
        ("womens", "Women"),
        ("home_visit", "Home"),
    ]
    counts = [all_appointments.filter(service=k).count() for k, _ in service_map]
    max_count = max(counts) if counts and max(counts) > 0 else 1
    service_stats = []
    for (key, short), count in zip(service_map, counts):
        service_stats.append(
            {
                "short": short,
                "count": count,
                "bar_height": max(round(count / max_count * 120), 4)
                if count > 0
                else 4,
            }
        )

    # ── Performance rates ──
    completion_rate = pct(completed_appointments)
    retention_rate = min(completion_rate + 15, 100)
    total_tasks = DailyTask.objects.count()
    completed_tasks = DailyTask.objects.filter(status="completed").count()
    task_completion_rate = (
        round(completed_tasks / total_tasks * 100) if total_tasks else 0
    )
    total_notes = SessionNote.objects.count()
    notes_rate = (
        round(total_notes / completed_appointments * 100)
        if completed_appointments
        else 0
    )

    return render(
        request,
        "reports_analytics.html",
        {
            "total_patients": User.objects.filter(
                is_superuser=False, is_staff=False
            ).count(),
            "total_appointments": total_appointments,
            "completed_appointments": completed_appointments,
            "confirmed_appointments": confirmed_appointments,
            "pending_appointments": pending_appointments,
            "cancelled_appointments": cancelled_appointments,
            "confirmed_pct": confirmed_pct,
            "completed_pct": completed_pct,
            "pending_pct": pending_pct,
            "service_stats": service_stats,
            "completion_rate": completion_rate,
            "retention_rate": retention_rate,
            "task_completion_rate": task_completion_rate,
            "notes_rate": min(notes_rate, 100),
            "total_staff": StaffProfile.objects.count(),
            "total_notes": total_notes,
            "total_leaves": LeaveApplication.objects.count(),
            "total_tasks": total_tasks,
            "recent_patients": User.objects.filter(
                is_superuser=False, is_staff=False
            ).order_by("-date_joined")[:6],
            "recent_appointments": all_appointments.order_by("-created_at")[:8],
            "staff_list": StaffProfile.objects.all().select_related("user"),
        },
    )


# ════════════════════════════════════════════════════════════
# EXPORT VIEWS — Excel & PDF
# ════════════════════════════════════════════════════════════

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER


# ─── PATIENTS EXPORT ─────────────────────────────────────────


@login_required
def export_patients_excel(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "#",
        "Full Name",
        "Username",
        "Email",
        "Phone",
        "Gender",
        "Date of Birth",
        "Blood Group",
        "Address",
        "Appointments",
        "Status",
        "Joined Date",
    ]
    ws.append(headers)

    for col_idx, col in enumerate(ws[1], 1):
        col.font = header_font
        col.fill = header_fill
        col.alignment = header_align
        col.border = border

    patients = User.objects.filter(is_superuser=False, is_staff=False).order_by(
        "-date_joined"
    )
    for i, p in enumerate(patients, 1):
        try:
            prof = p.profile
            phone = prof.phone_number or "—"
            gender = prof.gender or "—"
            dob = str(prof.date_of_birth) if prof.date_of_birth else "—"
            blood = prof.blood_group or "—"
            address = prof.address or "—"
        except Exception:
            phone = gender = dob = blood = address = "—"
        appt_count = Appointment.objects.filter(patient=p).count()
        status = "Active" if p.is_active else "Deactivated"
        row = [
            i,
            p.get_full_name() or p.username,
            p.username,
            p.email,
            phone,
            gender,
            dob,
            blood,
            address,
            appt_count,
            status,
            p.date_joined.strftime("%d %b %Y"),
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Column widths
    col_widths = [5, 22, 18, 28, 16, 10, 14, 12, 30, 14, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="patients_list.xlsx"'
    return response


@login_required
def export_patients_pdf(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1a1a1a"),
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
        alignment=TA_CENTER,
    )
    elements.append(Paragraph("Dr. Dhvani Patalia Physio-Rehab", title_style))
    elements.append(
        Paragraph(
            f"Patient List — Generated on {timezone.now().strftime('%d %b %Y %H:%M')}",
            sub_style,
        )
    )

    patients = User.objects.filter(is_superuser=False, is_staff=False).order_by(
        "-date_joined"
    )
    data = [
        [
            "#",
            "Full Name",
            "Email",
            "Phone",
            "Gender",
            "Blood",
            "Appointments",
            "Status",
            "Joined",
        ]
    ]
    for i, p in enumerate(patients, 1):
        try:
            prof = p.profile
            phone = prof.phone_number or "—"
            gender = prof.gender or "—"
            blood = prof.blood_group or "—"
        except Exception:
            phone = gender = blood = "—"
        appt_count = Appointment.objects.filter(patient=p).count()
        status = "Active" if p.is_active else "Deactivated"
        data.append(
            [
                str(i),
                p.get_full_name() or p.username,
                p.email or "—",
                phone,
                gender,
                blood,
                str(appt_count),
                status,
                p.date_joined.strftime("%d %b %Y"),
            ]
        )

    col_widths = [
        0.8 * cm,
        4 * cm,
        5.5 * cm,
        3 * cm,
        2 * cm,
        1.8 * cm,
        2.5 * cm,
        2.5 * cm,
        3 * cm,
    ]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a1a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f9f9f9")],
                ),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ("ROWHEIGHT", (0, 0), (-1, -1), 18),
            ]
        )
    )
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="patients_list.pdf"'
    return response


# ─── APPOINTMENTS EXPORT ─────────────────────────────────────


@login_required
def export_appointments_excel(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "#",
        "Patient Name",
        "Email",
        "Phone",
        "Service",
        "Date",
        "Time",
        "Status",
        "Notes",
        "Booked On",
    ]
    ws.append(headers)
    for col in ws[1]:
        col.font = header_font
        col.fill = header_fill
        col.alignment = header_align
        col.border = border

    appts = Appointment.objects.all().order_by("-created_at")
    for i, a in enumerate(appts, 1):
        row = [
            i,
            a.name,
            a.email,
            a.phone,
            a.service,
            str(a.date),
            a.time,
            a.status.upper(),
            a.notes or "—",
            a.created_at.strftime("%d %b %Y"),
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    col_widths = [5, 22, 28, 16, 20, 14, 10, 14, 30, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="appointments.xlsx"'
    return response


@login_required
def export_appointments_pdf(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1a1a1a"),
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
        alignment=TA_CENTER,
    )
    elements.append(Paragraph("Dr. Dhvani Patalia Physio-Rehab", title_style))
    elements.append(
        Paragraph(
            f"Appointments Report — Generated on {timezone.now().strftime('%d %b %Y %H:%M')}",
            sub_style,
        )
    )

    appts = Appointment.objects.all().order_by("-created_at")
    data = [
        ["#", "Patient Name", "Phone", "Service", "Date", "Time", "Status", "Booked On"]
    ]
    for i, a in enumerate(appts, 1):
        data.append(
            [
                str(i),
                a.name,
                a.phone,
                a.service,
                str(a.date),
                a.time,
                a.status.upper(),
                a.created_at.strftime("%d %b %Y"),
            ]
        )

    col_widths = [
        0.8 * cm,
        4.5 * cm,
        3 * cm,
        4.5 * cm,
        3 * cm,
        2.5 * cm,
        3 * cm,
        3 * cm,
    ]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a1a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f9f9f9")],
                ),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ("ROWHEIGHT", (0, 0), (-1, -1), 18),
            ]
        )
    )
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="appointments.pdf"'
    return response


# ─── ANALYTICS REPORT PDF ────────────────────────────────────


@login_required
def export_analytics_pdf(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1a1a1a"),
        spaceAfter=4,
        alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER,
    )
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=colors.HexColor("#f5c518"),
        spaceBefore=14,
        spaceAfter=6,
    )

    elements.append(Paragraph("Dr. Dhvani Patalia Physio-Rehab", title_style))
    elements.append(
        Paragraph(
            f"Analytics Report — {timezone.now().strftime('%d %b %Y')}", sub_style
        )
    )

    total_patients = User.objects.filter(is_superuser=False, is_staff=False).count()
    active_patients = User.objects.filter(
        is_superuser=False, is_staff=False, is_active=True
    ).count()
    total_appts = Appointment.objects.count()
    completed = Appointment.objects.filter(status="completed").count()
    confirmed = Appointment.objects.filter(status="confirmed").count()
    pending = Appointment.objects.filter(status="pending").count()
    cancelled = Appointment.objects.filter(status="cancelled").count()
    total_staff = StaffProfile.objects.count()

    elements.append(Paragraph("Summary Statistics", section_style))
    summary_data = [
        ["Metric", "Count"],
        ["Total Patients", str(total_patients)],
        ["Active Patients", str(active_patients)],
        ["Deactivated Patients", str(total_patients - active_patients)],
        ["Total Staff", str(total_staff)],
        ["Total Appointments", str(total_appts)],
        ["Completed Appointments", str(completed)],
        ["Confirmed Appointments", str(confirmed)],
        ["Pending Appointments", str(pending)],
        ["Cancelled Appointments", str(cancelled)],
    ]
    t = Table(summary_data, colWidths=[10 * cm, 5 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a1a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f9f9f9")],
                ),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ("ROWHEIGHT", (0, 0), (-1, -1), 20),
            ]
        )
    )
    elements.append(t)

    elements.append(Paragraph("Service Breakdown", section_style))
    from django.db.models import Count

    service_stats = (
        Appointment.objects.values("service")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    svc_data = [["Service", "Total Appointments"]]
    for s in service_stats:
        svc_data.append([s["service"].replace("_", " ").title(), str(s["count"])])
    if len(svc_data) > 1:
        t2 = Table(svc_data, colWidths=[10 * cm, 5 * cm])
        t2.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a1a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f9f9f9")],
                    ),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                    ("ROWHEIGHT", (0, 0), (-1, -1), 20),
                ]
            )
        )
        elements.append(t2)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="analytics_report.pdf"'
    return response


# ════════════════════════════════════════════════════════════
# SUPER ADMIN PANEL — Platform-level multi-clinic management
# ════════════════════════════════════════════════════════════


@platform_superadmin_required
def super_admin_dashboard(request):
    from django.db.models import Count, Sum

    total_hospitals = Hospital.objects.count()
    active_hospitals = Hospital.objects.filter(is_active=True).count()
    total_patients = User.objects.filter(is_superuser=False, is_staff=False).count()
    total_staff = StaffProfile.objects.count()
    total_appointments = Appointment.objects.count()
    open_tickets = SupportTicket.objects.filter(status="open").count()
    active_subs = HospitalSubscription.objects.filter(status="active").count()
    trial_subs = HospitalSubscription.objects.filter(status="trial").count()
    recent_hospitals = Hospital.objects.order_by("-created_at")[:6]
    recent_tickets = SupportTicket.objects.select_related("hospital").order_by(
        "-created_at"
    )[:5]
    recent_payments = PaymentRecord.objects.select_related(
        "patient", "appointment"
    ).order_by("-created_at")[:15]
    total_revenue = (
        PaymentRecord.objects.filter(status="paid").aggregate(t=Sum("amount"))["t"] or 0
    )
    pending_payments = PaymentRecord.objects.filter(status="pending").count()
    return render(
        request,
        "super_admin_dashboard.html",
        {
            "total_hospitals": total_hospitals,
            "active_hospitals": active_hospitals,
            "total_patients": total_patients,
            "total_staff": total_staff,
            "total_appointments": total_appointments,
            "open_tickets": open_tickets,
            "active_subs": active_subs,
            "trial_subs": trial_subs,
            "recent_hospitals": recent_hospitals,
            "recent_tickets": recent_tickets,
            "recent_payments": recent_payments,
            "total_revenue": total_revenue,
            "pending_payments": pending_payments,
        },
    )


@platform_superadmin_required
def super_admin_hospitals(request):
    hospitals = Hospital.objects.prefetch_related("subscription", "admins__user").all()
    return render(request, "super_admin_hospitals.html", {"hospitals": hospitals})


@platform_superadmin_required
def super_admin_add_hospital(request):
    plans = SubscriptionPlan.objects.filter(is_active=True)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        city = request.POST.get("city", "").strip()
        address = request.POST.get("address", "").strip()
        phone = request.POST.get("phone", "").strip()
        email = request.POST.get("email", "").strip()
        admin_username = request.POST.get("admin_username", "").strip()
        admin_email = request.POST.get("admin_email", "").strip()
        admin_password = request.POST.get("admin_password", "").strip()
        admin_first = request.POST.get("admin_first", "").strip()
        admin_last = request.POST.get("admin_last", "").strip()
        plan_id = request.POST.get("plan", "")

        errors = []
        if not name:
            errors.append("Hospital name is required.")
        if not admin_username:
            errors.append("Admin username is required.")
        if not admin_email:
            errors.append("Admin email is required.")
        if not admin_password:
            errors.append("Admin password is required.")
        if User.objects.filter(username=admin_username).exists():
            errors.append(f'Username "{admin_username}" is already taken.')
        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, "super_admin_add_hospital.html", {"plans": plans})

        hospital = Hospital.objects.create(
            name=name, city=city, address=address, phone=phone, email=email
        )
        admin_user = User.objects.create_user(
            username=admin_username,
            email=admin_email,
            password=admin_password,
            first_name=admin_first,
            last_name=admin_last,
            is_superuser=True,
            is_staff=True,
        )
        ClinicAdmin.objects.create(user=admin_user, hospital=hospital)

        plan = None
        if plan_id:
            try:
                plan = SubscriptionPlan.objects.get(id=plan_id)
            except SubscriptionPlan.DoesNotExist:
                pass
        HospitalSubscription.objects.create(
            hospital=hospital, plan=plan, status="trial"
        )

        messages.success(
            request, f'✅ Hospital "{name}" created with admin "{admin_username}".'
        )
        return redirect("super_admin_hospitals")

    return render(request, "super_admin_add_hospital.html", {"plans": plans})


@platform_superadmin_required
def super_admin_edit_hospital(request, hospital_id):
    hospital = get_object_or_404(Hospital, id=hospital_id)
    plans = SubscriptionPlan.objects.filter(is_active=True)
    try:
        subscription = hospital.subscription
    except HospitalSubscription.DoesNotExist:
        subscription = None

    if request.method == "POST":
        hospital.name = request.POST.get("name", hospital.name).strip()
        hospital.city = request.POST.get("city", "").strip()
        hospital.address = request.POST.get("address", "").strip()
        hospital.phone = request.POST.get("phone", "").strip()
        hospital.email = request.POST.get("email", "").strip()
        hospital.is_active = request.POST.get("is_active") == "on"
        hospital.save()

        plan_id = request.POST.get("plan", "")
        sub_status = request.POST.get("sub_status", "trial")
        expires_raw = request.POST.get("expires_at", "")
        expires_at = expires_raw if expires_raw else None

        if subscription:
            if plan_id:
                try:
                    subscription.plan = SubscriptionPlan.objects.get(id=plan_id)
                except SubscriptionPlan.DoesNotExist:
                    pass
            subscription.status = sub_status
            subscription.expires_at = expires_at
            subscription.save()
        else:
            plan = None
            if plan_id:
                try:
                    plan = SubscriptionPlan.objects.get(id=plan_id)
                except SubscriptionPlan.DoesNotExist:
                    pass
            HospitalSubscription.objects.create(
                hospital=hospital, plan=plan, status=sub_status, expires_at=expires_at
            )

        messages.success(request, f'✅ Hospital "{hospital.name}" updated.')
        return redirect("super_admin_hospitals")

    return render(
        request,
        "super_admin_edit_hospital.html",
        {
            "hospital": hospital,
            "plans": plans,
            "subscription": subscription,
        },
    )


@platform_superadmin_required
def super_admin_delete_hospital(request, hospital_id):
    hospital = get_object_or_404(Hospital, id=hospital_id)
    if request.method == "POST":
        name = hospital.name
        hospital.delete()
        messages.success(request, f'Hospital "{name}" deleted.')
    return redirect("super_admin_hospitals")


@platform_superadmin_required
def super_admin_subscriptions(request):
    from django.db.models import Sum, Count

    subscriptions = HospitalSubscription.objects.select_related(
        "hospital", "plan"
    ).all()
    plans = SubscriptionPlan.objects.all()
    # Pending payments needing verification
    pending_payments = (
        ClinicSubscriptionPayment.objects.select_related("hospital", "plan")
        .filter(status="pending")
        .order_by("-created_at")
    )
    # Per-hospital payment summary
    hospital_payment_summary = {}
    for csp in ClinicSubscriptionPayment.objects.select_related(
        "hospital", "plan"
    ).order_by("-created_at"):
        hid = csp.hospital_id
        if hid not in hospital_payment_summary:
            hospital_payment_summary[hid] = {
                "hospital": csp.hospital,
                "payments": [],
                "total_paid": 0,
                "has_pending": False,
            }
        hospital_payment_summary[hid]["payments"].append(csp)
        if csp.status == "paid":
            hospital_payment_summary[hid]["total_paid"] += float(csp.amount)
        if csp.status == "pending":
            hospital_payment_summary[hid]["has_pending"] = True

    # Save plan changes
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "save_plan":
            plan_name = request.POST.get("plan_name", "").strip()
            price_monthly = request.POST.get("price_monthly", "0")
            max_staff = request.POST.get("max_staff", "5")
            max_patients = request.POST.get("max_patients", "500")
            features = request.POST.get("features", "")
            plan_id = request.POST.get("plan_id", "")
            if plan_id:
                try:
                    p = SubscriptionPlan.objects.get(id=plan_id)
                    p.price_monthly = price_monthly
                    p.max_staff = max_staff
                    p.max_patients = max_patients
                    p.features = features
                    p.save()
                    messages.success(request, "Plan updated.")
                except SubscriptionPlan.DoesNotExist:
                    pass
            else:
                obj, created = SubscriptionPlan.objects.get_or_create(
                    name=plan_name,
                    defaults={
                        "price_monthly": price_monthly,
                        "max_staff": max_staff,
                        "max_patients": max_patients,
                        "features": features,
                    },
                )
                if not created:
                    obj.price_monthly = price_monthly
                    obj.max_staff = max_staff
                    obj.max_patients = max_patients
                    obj.features = features
                    obj.is_active = True
                    obj.save()
                    messages.success(
                        request, f"{obj.get_name_display()} plan updated successfully."
                    )
                else:
                    messages.success(
                        request, f"{obj.get_name_display()} plan created successfully."
                    )
        return redirect("super_admin_subscriptions")

    unpaid_clinics = []
    for sub in subscriptions:
        if sub.status in ("trial", "expired") or sub.is_expired:
            hid = sub.hospital_id
            summary = hospital_payment_summary.get(hid, {})
            if not summary.get("total_paid", 0):
                unpaid_clinics.append(sub.hospital)

    return render(
        request,
        "super_admin_subscriptions.html",
        {
            "subscriptions": subscriptions,
            "plans": plans,
            "pending_payments": pending_payments,
            "hospital_payment_summary": hospital_payment_summary,
            "unpaid_clinics": unpaid_clinics,
        },
    )


@platform_superadmin_required
def super_admin_confirm_sub_payment(request, payment_id):
    payment = get_object_or_404(ClinicSubscriptionPayment, id=payment_id)
    from dateutil.relativedelta import relativedelta

    payment.status = "paid"
    payment.paid_at = timezone.now()
    payment.save()
    # Activate the hospital subscription
    hospital = payment.hospital
    try:
        sub = hospital.subscription
    except HospitalSubscription.DoesNotExist:
        sub = HospitalSubscription(hospital=hospital)
    sub.plan = payment.plan
    sub.status = "active"
    sub.started_at = timezone.now().date()
    sub.expires_at = timezone.now().date() + relativedelta(
        months=payment.duration_months
    )
    sub.save()
    messages.success(
        request,
        f"✅ Payment confirmed and {payment.hospital.name} subscription activated until {sub.expires_at}.",
    )
    return redirect("super_admin_subscriptions")


@platform_superadmin_required
def super_admin_reject_sub_payment(request, payment_id):
    payment = get_object_or_404(ClinicSubscriptionPayment, id=payment_id)
    payment.status = "failed"
    payment.save()
    messages.warning(
        request, f"❌ Payment for {payment.hospital.name} marked as failed."
    )
    return redirect("super_admin_subscriptions")


@platform_superadmin_required
def super_admin_hospital_payments(request, hospital_id):
    hospital = get_object_or_404(Hospital, id=hospital_id)
    payments = (
        ClinicSubscriptionPayment.objects.filter(hospital=hospital)
        .select_related("plan")
        .order_by("-created_at")
    )
    try:
        current_sub = hospital.subscription
    except HospitalSubscription.DoesNotExist:
        current_sub = None
    return render(
        request,
        "super_admin_hospital_payments.html",
        {
            "hospital": hospital,
            "payments": payments,
            "current_sub": current_sub,
        },
    )


@platform_superadmin_required
def super_admin_analytics(request):
    from django.db.models import Count

    hospitals = Hospital.objects.annotate(
        appt_count=Count("appointments", distinct=True),
        staff_count=Count("staff", distinct=True),
    ).order_by("-appt_count")
    total_hospitals = Hospital.objects.count()
    active_hospitals = Hospital.objects.filter(is_active=True).count()
    total_appointments = Appointment.objects.count()
    total_patients = User.objects.filter(is_superuser=False, is_staff=False).count()
    total_staff = StaffProfile.objects.count()
    active_subs = HospitalSubscription.objects.filter(status="active").count()
    trial_subs = HospitalSubscription.objects.filter(status="trial").count()
    expired_subs = HospitalSubscription.objects.filter(status="expired").count()
    return render(
        request,
        "super_admin_analytics.html",
        {
            "hospitals": hospitals,
            "total_hospitals": total_hospitals,
            "active_hospitals": active_hospitals,
            "total_appointments": total_appointments,
            "total_patients": total_patients,
            "total_staff": total_staff,
            "active_subs": active_subs,
            "trial_subs": trial_subs,
            "expired_subs": expired_subs,
        },
    )


@platform_superadmin_required
def super_admin_support(request):
    status_filter = request.GET.get("status", "")
    tickets = SupportTicket.objects.select_related("hospital", "submitted_by").all()
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    open_count = SupportTicket.objects.filter(status="open").count()
    in_progress_count = SupportTicket.objects.filter(status="in_progress").count()
    resolved_count = SupportTicket.objects.filter(status="resolved").count()
    return render(
        request,
        "super_admin_support.html",
        {
            "tickets": tickets,
            "status_filter": status_filter,
            "open_count": open_count,
            "in_progress_count": in_progress_count,
            "resolved_count": resolved_count,
        },
    )


@platform_superadmin_required
def super_admin_support_reply(request, ticket_id):
    ticket = get_object_or_404(SupportTicket, id=ticket_id)
    replies = ticket.replies.select_related("replier").all()
    hospitals = Hospital.objects.filter(is_active=True)
    if request.method == "POST":
        msg = request.POST.get("message", "").strip()
        new_status = request.POST.get("status", ticket.status)
        if msg:
            SupportReply.objects.create(
                ticket=ticket, replier=request.user, message=msg
            )
        ticket.status = new_status
        ticket.save()
        messages.success(request, "✅ Reply sent.")
        return redirect("super_admin_support_reply", ticket_id=ticket_id)
    return render(
        request,
        "super_admin_support_reply.html",
        {
            "ticket": ticket,
            "replies": replies,
            "hospitals": hospitals,
        },
    )


# ── Clinic admin can submit a support ticket ──────────────────
@login_required
def submit_support_ticket(request):
    if not _is_clinic_admin(request.user):
        return redirect("home")
    hospital = None
    if hasattr(request.user, "clinic_admin_profile"):
        hospital = request.user.clinic_admin_profile.hospital
    if request.method == "POST":
        subject = request.POST.get("subject", "").strip()
        msg = request.POST.get("message", "").strip()
        if subject and msg:
            SupportTicket.objects.create(
                hospital=hospital,
                submitted_by=request.user,
                subject=subject,
                message=msg,
            )
            messages.success(request, "✅ Support ticket submitted successfully.")
        return redirect("admin_dashboard")
    return redirect("admin_dashboard")


# ─── SUBSCRIPTION PAGE ────────────────────────────────────────


@platform_superadmin_required
def super_admin_all_payments(request):
    from django.db.models import Sum
    status_filter = request.GET.get("status", "")
    method_filter = request.GET.get("method", "")
    all_payments = PaymentRecord.objects.select_related("patient", "appointment").order_by("-created_at")
    if status_filter:
        all_payments = all_payments.filter(status=status_filter)
    if method_filter:
        all_payments = all_payments.filter(method=method_filter)
    total_paid = PaymentRecord.objects.filter(status="paid").aggregate(t=Sum("amount"))["t"] or 0
    total_pending = PaymentRecord.objects.filter(status="pending").aggregate(t=Sum("amount"))["t"] or 0
    pending_count = PaymentRecord.objects.filter(status="pending").count()
    paid_count = PaymentRecord.objects.filter(status="paid").count()
    failed_count = PaymentRecord.objects.filter(status="failed").count()
    return render(request, "super_admin_all_payments.html", {
        "payments": all_payments,
        "status_filter": status_filter,
        "method_filter": method_filter,
        "total_paid": total_paid,
        "total_pending": total_pending,
        "pending_count": pending_count,
        "paid_count": paid_count,
        "failed_count": failed_count,
    })


@platform_superadmin_required
def export_payments_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    status_filter = request.GET.get("status", "")
    method_filter = request.GET.get("method", "")
    qs = PaymentRecord.objects.select_related("patient", "appointment").order_by("-created_at")
    if status_filter:
        qs = qs.filter(status=status_filter)
    if method_filter:
        qs = qs.filter(method=method_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Payments"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1a1a2e")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Date", "Time", "Patient Name", "Patient Email", "Appointment Date", "Amount (₹)", "Method", "Transaction ID / Ref", "Notes", "Status"]
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    col_widths = [5, 14, 10, 22, 28, 16, 14, 14, 26, 30, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    alt_fill = PatternFill("solid", fgColor="F5F5FF")
    for i, pay in enumerate(qs, 1):
        row = [
            i,
            pay.created_at.strftime("%d %b %Y"),
            pay.created_at.strftime("%I:%M %p"),
            pay.patient.get_full_name() or pay.patient.username,
            pay.patient.email or "—",
            str(pay.appointment.date) if pay.appointment else "General",
            float(pay.amount),
            pay.get_method_display(),
            pay.transaction_id or "—",
            pay.notes or "—",
            pay.get_status_display(),
        ]
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i + 1]:
                cell.fill = alt_fill
        for cell in ws[i + 1]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="all_payments.xlsx"'
    wb.save(resp)
    return resp


@platform_superadmin_required
def export_payments_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from django.http import HttpResponse
    from django.utils import timezone as tz

    status_filter = request.GET.get("status", "")
    method_filter = request.GET.get("method", "")
    qs = PaymentRecord.objects.select_related("patient", "appointment").order_by("-created_at")
    if status_filter:
        qs = qs.filter(status=status_filter)
    if method_filter:
        qs = qs.filter(method=method_filter)

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="all_payments.pdf"'
    doc = SimpleDocTemplate(resp, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("title", fontSize=16, fontName="Helvetica-Bold", textColor=colors.HexColor("#1a1a2e"), spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle("sub", fontSize=9, fontName="Helvetica", textColor=colors.grey, spaceAfter=12, alignment=TA_CENTER)
    elements.append(Paragraph("Patient Payment Transactions", title_style))
    elements.append(Paragraph(f"Generated: {tz.now().strftime('%d %b %Y, %I:%M %p')}   |   Records: {qs.count()}", sub_style))

    headers = ["#", "Date", "Patient", "Amount", "Method", "Txn / Ref", "Status"]
    data = [headers]
    for i, pay in enumerate(qs, 1):
        data.append([
            str(i),
            pay.created_at.strftime("%d %b %Y"),
            (pay.patient.get_full_name() or pay.patient.username)[:22],
            f"Rs.{pay.amount}",
            pay.get_method_display(),
            (pay.transaction_id or "—")[:20],
            pay.get_status_display(),
        ])

    col_widths_pdf = [1.0*cm, 2.8*cm, 5.5*cm, 2.5*cm, 2.8*cm, 5.0*cm, 2.5*cm]
    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5FF")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWHEIGHT", (0, 0), (-1, 0), 20),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(tbl)
    doc.build(elements)
    return resp


@platform_superadmin_required
@require_POST
def send_subscription_reminder(request, hospital_id):
    from .models import ClinicAdmin as ClinicAdminModel
    hospital = get_object_or_404(Hospital, id=hospital_id)
    admin_users = list(ClinicAdminModel.objects.filter(hospital=hospital).select_related("user"))
    if not admin_users:
        messages.warning(request, f"No admin found for {hospital.name}.")
        return redirect("super_admin_subscriptions")
    for ca in admin_users:
        Notification.objects.create(
            recipient=ca.user,
            message=f"⚠️ Reminder: Your clinic subscription is due. Please visit the subscription page to renew or upgrade your plan to continue using all features.",
            link="/subscription/",
        )
    messages.success(request, f"✅ Subscription reminder sent to {len(admin_users)} admin(s) of {hospital.name}.")
    return redirect("super_admin_subscriptions")


@login_required
def subscription_page(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    reason = request.GET.get("reason", "")
    if reason == "expired":
        messages.error(request, "❌ Your subscription has expired. Please renew your plan to restore full access.")
    elif reason == "trial":
        messages.warning(request, "🔒 This feature is locked on the trial plan. Upgrade to unlock all premium features.")
    plans = SubscriptionPlan.objects.filter(is_active=True).order_by("price_monthly")
    hospital = None
    current_sub = None
    recent_payments = []
    upi_id = getattr(settings, "CLINIC_UPI_ID", "dhvanipatalia@upi")

    if hasattr(request.user, "clinic_admin_profile"):
        hospital = request.user.clinic_admin_profile.hospital
        try:
            current_sub = hospital.subscription
        except Exception:
            pass
        recent_payments = ClinicSubscriptionPayment.objects.filter(
            hospital=hospital
        ).order_by("-created_at")[:10]

    if request.method == "POST":
        if not hospital:
            messages.error(request, "No clinic linked to your account.")
            return redirect("subscription_page")
        plan_id = request.POST.get("plan_id")
        method = request.POST.get("method", "upi")
        txn_id = request.POST.get("transaction_id", "").strip()
        duration = int(request.POST.get("duration_months", 1))
        notes = request.POST.get("notes", "").strip()
        try:
            plan = SubscriptionPlan.objects.get(id=plan_id)
        except SubscriptionPlan.DoesNotExist:
            messages.error(request, "Invalid plan selected.")
            return redirect("subscription_page")

        amount = plan.price_monthly * duration
        ClinicSubscriptionPayment.objects.create(
            hospital=hospital,
            plan=plan,
            amount=amount,
            method=method,
            status="pending",
            transaction_id=txn_id,
            duration_months=duration,
            notes=notes,
        )
        # Update the hospital subscription plan to the requested one
        try:
            sub = hospital.subscription
            sub.plan = plan
            sub.save()
        except Exception:
            HospitalSubscription.objects.create(
                hospital=hospital, plan=plan, status="trial"
            )

        messages.success(
            request,
            f"✅ Payment submitted for {plan.get_name_display()} plan! The platform admin will verify and activate your subscription shortly.",
        )
        return redirect("subscription_page")

    return render(
        request,
        "subscription_page.html",
        {
            "plans": plans,
            "hospital": hospital,
            "current_sub": current_sub,
            "recent_payments": recent_payments,
            "upi_id": upi_id,
        },
    )


# ─── PROMO MANAGEMENT ────────────────────────────────────────


@login_required
def admin_promos(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    promos = ClinicPromo.objects.all()
    return render(request, "admin_promos.html", {"promos": promos})


@login_required
def admin_promo_add(request):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    if request.method == "POST":
        ClinicPromo.objects.create(
            title=request.POST.get("title", "").strip(),
            message=request.POST.get("message", "").strip(),
            start_date=request.POST.get("start_date") or timezone.now().date(),
            end_date=request.POST.get("end_date") or None,
            is_active=True,
        )
        messages.success(request, "✅ Promo created.")
        return redirect("admin_promos")
    return render(request, "admin_promo_add.html")


@login_required
def admin_promo_delete(request, promo_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    promo = get_object_or_404(ClinicPromo, id=promo_id)
    promo.delete()
    messages.success(request, "🗑️ Promo deleted.")
    return redirect("admin_promos")


@login_required
def dismiss_promo(request):
    if request.method == "POST":
        promo_id = request.POST.get("promo_id")
        dismissed = request.session.get("dismissed_promos", [])
        if promo_id and promo_id not in dismissed:
            dismissed.append(promo_id)
            request.session["dismissed_promos"] = dismissed
    return JsonResponse({"ok": True})


# ─── EXERCISE REMINDER ────────────────────────────────────────


@login_required
def send_exercise_reminder(request, patient_id):
    if not request.user.is_superuser:
        return redirect("client_dashboard")
    patient = get_object_or_404(User, id=patient_id)
    last_note = (
        SessionNote.objects.filter(patient=patient).order_by("-created_at").first()
    )
    next_info = last_note.next_session if last_note else "your scheduled session"
    from .notifications import notify

    notify(
        recipient=patient,
        message=f"🏃 Reminder from Dr. Dhvani: Please complete your exercises before {next_info}. Stay consistent for faster recovery!",
        link="/my-appointments/",
    )
    if patient.email:
        from .email_utils import send_clinic_email

        send_clinic_email(
            subject="Exercise Reminder from Dr. Dhvani Patalia",
            message_text=f"""Dear {patient.get_full_name() or patient.username},

This is a friendly reminder from Dr. Dhvani Patalia's PhysioRehab Clinic.

Please ensure you complete your prescribed exercises before your next session: {next_info}

Consistency is key to a faster recovery. If you have any questions, feel free to chat with us.

Dr. Dhvani Patalia — PhysioRehab Clinic
""",
            recipient_list=[patient.email],
        )
    messages.success(
        request,
        f"✅ Exercise reminder sent to {patient.get_full_name() or patient.username}.",
    )
    return redirect("admin_patients")
